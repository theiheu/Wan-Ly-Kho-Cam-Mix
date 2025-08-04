#!/usr/bin/env python3
"""
Daily Feed Excel Export Service - Xuất Excel báo cáo tiêu thụ cám hàng ngày
Tích hợp với hệ thống cache và comprehensive reporting
"""

import os
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

# Import services
try:
    from src.services.daily_report_calculator import daily_report_calculator
    from src.services.cached_report_viewer import cached_report_viewer
    from src.utils.user_preferences import user_preferences_manager
except ImportError:
    from services.daily_report_calculator import daily_report_calculator
    from services.cached_report_viewer import cached_report_viewer
    from utils.user_preferences import user_preferences_manager

class DailyFeedExcelExporter:
    """Xuất Excel báo cáo tiêu thụ cám hàng ngày"""

    def __init__(self):
        """Khởi tạo exporter"""
        self.calculator = daily_report_calculator
        self.viewer = cached_report_viewer
        self.preferences = user_preferences_manager

        # Thiết lập styles
        self.setup_excel_styles()

    def setup_excel_styles(self):
        """Thiết lập các style cho Excel"""
        # Font styles
        self.title_font = Font(name='Arial', size=16, bold=True, color='1F4E79')
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.subheader_font = Font(name='Arial', size=11, bold=True, color='1F4E79')
        self.normal_font = Font(name='Arial', size=10, color='000000')
        self.number_font = Font(name='Arial', size=10, color='000000')

        # Fill styles
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        self.alternate_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        self.summary_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')

        # Border styles
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        # Alignment styles
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')

    def create_workbook(self, report_date: str) -> Workbook:
        """Tạo workbook mới"""
        wb = Workbook()

        # Xóa worksheet mặc định
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        return wb

    def format_worksheet_header(self, ws: Worksheet, title: str, report_date: str, start_row: int = 1) -> int:
        """Định dạng header cho worksheet"""
        # Tiêu đề chính
        ws.cell(row=start_row, column=1, value=title)
        title_cell = ws.cell(row=start_row, column=1)
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment

        # Ngày báo cáo
        display_date = self._format_display_date(report_date)
        ws.cell(row=start_row + 1, column=1, value=f"Ngày: {display_date}")
        date_cell = ws.cell(row=start_row + 1, column=1)
        date_cell.font = self.subheader_font
        date_cell.alignment = self.left_alignment

        # Thời gian tạo
        ws.cell(row=start_row + 2, column=1, value=f"Tạo lúc: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        time_cell = ws.cell(row=start_row + 2, column=1)
        time_cell.font = self.normal_font
        time_cell.alignment = self.left_alignment

        return start_row + 4  # Trả về hàng tiếp theo

    def format_data_table(self, ws: Worksheet, df: pd.DataFrame, start_row: int, start_col: int = 1,
                         table_title: str = None) -> int:
        """Định dạng bảng dữ liệu"""
        current_row = start_row

        # Thêm tiêu đề bảng nếu có
        if table_title:
            ws.cell(row=current_row, column=start_col, value=table_title)
            title_cell = ws.cell(row=current_row, column=start_col)
            title_cell.font = self.subheader_font
            title_cell.fill = self.subheader_fill
            title_cell.alignment = self.left_alignment
            current_row += 2

        if df.empty:
            ws.cell(row=current_row, column=start_col, value="Không có dữ liệu")
            return current_row + 1

        # Thêm dữ liệu
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.thin_border

                # Định dạng header
                if r_idx == current_row:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
                else:
                    cell.font = self.normal_font

                    # Định dạng số
                    if isinstance(value, (int, float)) and value != 0:
                        cell.number_format = '#,##0.0'
                        cell.alignment = self.right_alignment
                    else:
                        cell.alignment = self.left_alignment

                    # Màu xen kẽ
                    if (r_idx - current_row) % 2 == 0:
                        cell.fill = self.alternate_fill

        # Tự động điều chỉnh độ rộng cột
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

        return current_row + len(df) + 2

    def create_feed_consumption_worksheet(self, wb: Workbook, report_date: str, report_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet tiêu thụ cám"""
        ws = wb.create_sheet(title="Tiêu Thụ Cám")

        current_row = self.format_worksheet_header(ws, "BÁO CÁO TIÊU THỤ CÁM THEO TRẠI", report_date)

        # Lấy dữ liệu bảng tiêu thụ cám
        feed_table_data = self.viewer.get_feed_consumption_table(report_date)

        if feed_table_data:
            # Chuyển đổi sang DataFrame
            feed_df = pd.DataFrame(feed_table_data)

            # Đổi tên cột sang tiếng Việt
            column_mapping = {
                'area': 'Khu Vực',
                'farm': 'Trại',
                'morning': 'Ca Sáng (kg)',
                'afternoon': 'Ca Chiều (kg)',
                'farm_total': 'Tổng Trại (kg)',
                'area_total': 'Tổng Khu (kg)',
                'percentage_of_area': 'Tỷ Lệ (%)'
            }

            feed_df = feed_df.rename(columns=column_mapping)

            # Định dạng bảng
            current_row = self.format_data_table(ws, feed_df, current_row, table_title="CHI TIẾT TIÊU THỤ CÁM THEO TRẠI VÀ CA")

            # Thêm ghi chú về sự khác biệt
            ws.cell(row=current_row, column=1, value="Ghi chú: Đây là dữ liệu tiêu thụ cám theo trại, khác với tổng sản xuất cám.")
            note_cell = ws.cell(row=current_row, column=1)
            note_cell.font = self.normal_font
            note_cell.fill = self.summary_fill
            current_row += 2

        # Thêm thống kê tổng quan
        if 'feed_calculations' in report_data:
            feed_calc = report_data['feed_calculations']

            # Tổng theo khu vực
            if 'area_totals' in feed_calc and 'areas' in feed_calc['area_totals']:
                area_data = []
                for area, data in feed_calc['area_totals']['areas'].items():
                    area_data.append({
                        'Khu Vực': area,
                        'Tổng Tiêu Thụ (kg)': data.get('total', 0),
                        'Số Trại': len(data.get('farms', {}))
                    })

                if area_data:
                    area_df = pd.DataFrame(area_data)
                    current_row = self.format_data_table(ws, area_df, current_row, table_title="TỔNG TIÊU THỤ THEO KHU VỰC")

            # Top trại tiêu thụ nhiều nhất
            if 'farm_rankings' in feed_calc:
                top_farms = feed_calc['farm_rankings'][:10]  # Top 10
                if top_farms:
                    ranking_data = []
                    for farm in top_farms:
                        ranking_data.append({
                            'Hạng': farm.get('rank', 0),
                            'Khu Vực': farm.get('area', ''),
                            'Trại': farm.get('farm', ''),
                            'Tổng Tiêu Thụ (kg)': farm.get('total_consumption', 0)
                        })

                    ranking_df = pd.DataFrame(ranking_data)
                    current_row = self.format_data_table(ws, ranking_df, current_row, table_title="TOP 10 TRẠI TIÊU THỤ CÁM NHIỀU NHẤT")

        return ws

    def create_summary_worksheet(self, wb: Workbook, report_date: str, report_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet tổng quan"""
        ws = wb.create_sheet(title="Tổng Quan", index=0)  # Đặt làm sheet đầu tiên

        current_row = self.format_worksheet_header(ws, "TỔNG QUAN BÁO CÁO TIÊU THỤ CÁM", report_date)

        # Thông tin báo cáo
        metadata = report_data.get('metadata', {})
        summary = report_data.get('summary', {})
        efficiency = report_data.get('efficiency_metrics', {})

        # Thông tin cơ bản
        basic_info = [
            ['Ngày báo cáo', self._format_display_date(report_date)],
            ['Thời gian tính toán', f"{metadata.get('calculation_time_seconds', 0):.3f} giây"],
            ['Tổng số khu vực hoạt động', summary.get('active_areas', 0)],
            ['Tổng số trại hoạt động', summary.get('active_farms', 0)],
            ['Sử dụng cache', 'Có' if metadata.get('cached', False) else 'Không']
        ]

        basic_df = pd.DataFrame(basic_info, columns=['Thông Tin', 'Giá Trị'])
        current_row = self.format_data_table(ws, basic_df, current_row, table_title="THÔNG TIN BÁO CÁO")

        # Thống kê tiêu thụ (bao gồm cả dữ liệu mix từ raw data)
        raw_data = report_data.get('raw_data', {})
        total_mix_ingredients = raw_data.get('total_mix', 0)
        mix_ingredients_count = len(raw_data.get('mix_ingredients', {}))
        feed_ingredients_count = len(raw_data.get('feed_ingredients', {}))

        # Lấy thông tin tiêu thụ theo trại
        feed_usage_total = efficiency.get('feed_usage_total', 0)
        mix_usage_total = efficiency.get('mix_usage_total', 0)

        consumption_info = [
            ['Tổng sản xuất cám (kg)', efficiency.get('feed_total', 0)],
            ['Tổng sản xuất mix (kg)', efficiency.get('mix_total', 0)],
            ['Tổng sản xuất (kg)', efficiency.get('total_consumption', 0)],
            ['Tiêu thụ cám theo trại (kg)', feed_usage_total],
            ['Tiêu thụ mix theo trại (kg)', mix_usage_total],
            ['Tỷ lệ cám (%)', f"{efficiency.get('feed_percentage', 0):.1f}%"],
            ['Tỷ lệ mix (%)', f"{efficiency.get('mix_percentage', 0):.1f}%"],
            ['Tỷ lệ cám/mix', f"{efficiency.get('feed_to_mix_ratio', 0):.2f}" if efficiency.get('feed_to_mix_ratio', 0) != float('inf') else 'N/A'],
            ['Số loại nguyên liệu feed', feed_ingredients_count],
            ['Số loại nguyên liệu mix', mix_ingredients_count],
            ['Hiệu suất tiêu thụ cám (%)', f"{efficiency.get('feed_usage_percentage', 0):.1f}%"],
            ['Hiệu suất tiêu thụ mix (%)', f"{efficiency.get('mix_usage_percentage', 0):.1f}%"]
        ]

        consumption_df = pd.DataFrame(consumption_info, columns=['Chỉ Số', 'Giá Trị'])
        current_row = self.format_data_table(ws, consumption_df, current_row, table_title="THỐNG KÊ TIÊU THỤ")

        # Trại tiêu thụ nhiều nhất
        top_farm = summary.get('top_consuming_farm')
        if top_farm:
            top_farm_info = [
                ['Trại tiêu thụ nhiều nhất', f"{top_farm.get('area', '')} - {top_farm.get('farm', '')}"],
                ['Lượng tiêu thụ (kg)', top_farm.get('total_consumption', 0)]
            ]

            top_farm_df = pd.DataFrame(top_farm_info, columns=['Thông Tin', 'Giá Trị'])
            current_row = self.format_data_table(ws, top_farm_df, current_row, table_title="TRẠI TIÊU THỤ HÀNG ĐẦU")

        return ws

    def _format_display_date(self, date_str: str) -> str:
        """Định dạng ngày hiển thị"""
        try:
            if len(date_str) == 8:
                year = date_str[:4]
                month = date_str[4:6]
                day = date_str[6:8]
                return f"{day}/{month}/{year}"
        except:
            pass
        return date_str

    def create_shift_analysis_worksheet(self, wb: Workbook, report_date: str, report_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet phân tích theo ca"""
        ws = wb.create_sheet(title="Phân Tích Ca")

        current_row = self.format_worksheet_header(ws, "PHÂN TÍCH TIÊU THỤ THEO CA", report_date)

        # Phân tích ca cho cám
        if 'feed_calculations' in report_data and 'shift_statistics' in report_data['feed_calculations']:
            shift_stats = report_data['feed_calculations']['shift_statistics']

            # Thống kê tổng theo ca
            shift_data = []
            for shift, total in shift_stats.get('totals', {}).items():
                average = shift_stats.get('averages', {}).get(shift, 0)
                count = shift_stats.get('counts', {}).get(shift, 0)

                shift_data.append({
                    'Ca': shift,
                    'Tổng Tiêu Thụ (kg)': total,
                    'Trung Bình/Trại (kg)': average,
                    'Số Trại Hoạt Động': count
                })

            if shift_data:
                shift_df = pd.DataFrame(shift_data)
                current_row = self.format_data_table(ws, shift_df, current_row, table_title="THỐNG KÊ TIÊU THỤ CÁM THEO CA")

        # Phân tích hiệu quả theo ca
        efficiency_data = []
        if 'feed_calculations' in report_data:
            feed_totals = report_data['feed_calculations'].get('shift_statistics', {}).get('totals', {})
            total_feed = sum(feed_totals.values())

            for shift, amount in feed_totals.items():
                percentage = (amount / total_feed * 100) if total_feed > 0 else 0
                efficiency_data.append({
                    'Ca': shift,
                    'Lượng Tiêu Thụ (kg)': amount,
                    'Tỷ Lệ (%)': percentage
                })

        if efficiency_data:
            efficiency_df = pd.DataFrame(efficiency_data)
            current_row = self.format_data_table(ws, efficiency_df, current_row, table_title="HIỆU QUẢ TIÊU THỤ THEO CA")

        return ws

    def create_area_analysis_worksheet(self, wb: Workbook, report_date: str, report_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet phân tích theo khu vực"""
        ws = wb.create_sheet(title="Phân Tích Khu Vực")

        current_row = self.format_worksheet_header(ws, "PHÂN TÍCH TIÊU THỤ THEO KHU VỰC", report_date)

        # Lấy tóm tắt khu vực
        area_summary = self.viewer.get_area_summary(report_date)

        if area_summary and area_summary.get('area_rankings'):
            # Bảng xếp hạng khu vực
            ranking_data = []
            for area_info in area_summary['area_rankings']:
                ranking_data.append({
                    'Hạng': area_info.get('rank', 0),
                    'Khu Vực': area_info.get('area', ''),
                    'Tổng Tiêu Thụ (kg)': area_info.get('total_consumption', 0),
                    'Tiêu Thụ Cám (kg)': area_info.get('feed_consumption', 0),
                    'Tiêu Thụ Mix (kg)': area_info.get('mix_consumption', 0),
                    'Tỷ Lệ Cám (%)': f"{area_info.get('feed_percentage', 0):.1f}%",
                    'Tỷ Lệ Mix (%)': f"{area_info.get('mix_percentage', 0):.1f}%"
                })

            ranking_df = pd.DataFrame(ranking_data)
            current_row = self.format_data_table(ws, ranking_df, current_row, table_title="XẾP HẠNG KHU VỰC THEO TIÊU THỤ")

        # Chi tiết theo khu vực từ feed calculations
        if 'feed_calculations' in report_data and 'area_totals' in report_data['feed_calculations']:
            area_details = []
            areas_data = report_data['feed_calculations']['area_totals'].get('areas', {})

            for area, data in areas_data.items():
                farms_count = len(data.get('farms', {}))
                total_consumption = data.get('total', 0)
                avg_per_farm = total_consumption / farms_count if farms_count > 0 else 0

                area_details.append({
                    'Khu Vực': area,
                    'Số Trại': farms_count,
                    'Tổng Tiêu Thụ (kg)': total_consumption,
                    'Trung Bình/Trại (kg)': avg_per_farm
                })

            if area_details:
                details_df = pd.DataFrame(area_details)
                current_row = self.format_data_table(ws, details_df, current_row, table_title="CHI TIẾT TIÊU THỤ CÁM THEO KHU VỰC")

        return ws

    def create_mix_consumption_worksheet(self, wb: Workbook, report_date: str, report_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet tiêu thụ mix"""
        ws = wb.create_sheet(title="Tiêu Thụ Mix")

        current_row = self.format_worksheet_header(ws, "BÁO CÁO TIÊU THỤ MIX HÀNG NGÀY", report_date)

        # Lấy dữ liệu mix từ raw data
        raw_data = report_data.get('raw_data', {})
        mix_ingredients = raw_data.get('mix_ingredients', {})
        total_mix = raw_data.get('total_mix', 0)

        # Tạo bảng nguyên liệu mix
        if mix_ingredients:
            mix_data = []
            total_ingredients = sum(mix_ingredients.values())

            for ingredient, amount in mix_ingredients.items():
                percentage = (amount / total_ingredients * 100) if total_ingredients > 0 else 0
                mix_data.append({
                    'Nguyên Liệu': ingredient,
                    'Khối Lượng (kg)': amount,
                    'Tỷ Lệ (%)': percentage,
                    'Đơn Giá (VND/kg)': 0,  # Có thể bổ sung sau
                    'Thành Tiền (VND)': 0   # Có thể bổ sung sau
                })

            # Sắp xếp theo khối lượng giảm dần
            mix_data.sort(key=lambda x: x['Khối Lượng (kg)'], reverse=True)

            mix_df = pd.DataFrame(mix_data)
            current_row = self.format_data_table(ws, mix_df, current_row, table_title="CHI TIẾT NGUYÊN LIỆU MIX SỬ DỤNG")

        # Thống kê tổng quan mix
        summary_data = [
            ['Tổng khối lượng mix (kg)', total_mix],
            ['Số loại nguyên liệu', len(mix_ingredients)],
            ['Nguyên liệu chính (>50kg)', len([x for x in mix_ingredients.values() if x > 50])],
            ['Nguyên liệu phụ (≤50kg)', len([x for x in mix_ingredients.values() if x <= 50])],
            ['Trung bình/nguyên liệu (kg)', sum(mix_ingredients.values()) / len(mix_ingredients) if mix_ingredients else 0]
        ]

        summary_df = pd.DataFrame(summary_data, columns=['Chỉ Số', 'Giá Trị'])
        current_row = self.format_data_table(ws, summary_df, current_row, table_title="THỐNG KÊ TỔNG QUAN MIX")

        # Top 10 nguyên liệu mix sử dụng nhiều nhất
        if mix_ingredients:
            top_ingredients = sorted(mix_ingredients.items(), key=lambda x: x[1], reverse=True)[:10]
            top_data = []

            for i, (ingredient, amount) in enumerate(top_ingredients, 1):
                percentage = (amount / sum(mix_ingredients.values()) * 100) if mix_ingredients else 0
                top_data.append({
                    'Hạng': i,
                    'Nguyên Liệu': ingredient,
                    'Khối Lượng (kg)': amount,
                    'Tỷ Lệ (%)': percentage
                })

            top_df = pd.DataFrame(top_data)
            current_row = self.format_data_table(ws, top_df, current_row, table_title="TOP 10 NGUYÊN LIỆU MIX SỬ DỤNG NHIỀU NHẤT")

        return ws

    def create_ingredients_comparison_worksheet(self, wb: Workbook, report_date: str, report_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet so sánh nguyên liệu feed vs mix"""
        ws = wb.create_sheet(title="So Sánh Nguyên Liệu")

        current_row = self.format_worksheet_header(ws, "SO SÁNH NGUYÊN LIỆU FEED VÀ MIX", report_date)

        # Lấy dữ liệu nguyên liệu
        raw_data = report_data.get('raw_data', {})
        feed_ingredients = raw_data.get('feed_ingredients', {})
        mix_ingredients = raw_data.get('mix_ingredients', {})

        # Tạo danh sách tất cả nguyên liệu
        all_ingredients = set(feed_ingredients.keys()) | set(mix_ingredients.keys())

        if all_ingredients:
            comparison_data = []

            for ingredient in sorted(all_ingredients):
                feed_amount = feed_ingredients.get(ingredient, 0)
                mix_amount = mix_ingredients.get(ingredient, 0)
                total_amount = feed_amount + mix_amount

                feed_percentage = (feed_amount / total_amount * 100) if total_amount > 0 else 0
                mix_percentage = (mix_amount / total_amount * 100) if total_amount > 0 else 0

                comparison_data.append({
                    'Nguyên Liệu': ingredient,
                    'Feed (kg)': feed_amount,
                    'Mix (kg)': mix_amount,
                    'Tổng (kg)': total_amount,
                    'Tỷ Lệ Feed (%)': feed_percentage,
                    'Tỷ Lệ Mix (%)': mix_percentage,
                    'Chênh Lệch (kg)': abs(feed_amount - mix_amount)
                })

            # Sắp xếp theo tổng khối lượng giảm dần
            comparison_data.sort(key=lambda x: x['Tổng (kg)'], reverse=True)

            comparison_df = pd.DataFrame(comparison_data)
            current_row = self.format_data_table(ws, comparison_df, current_row, table_title="SO SÁNH CHI TIẾT NGUYÊN LIỆU")

        # Thống kê tổng quan
        total_feed = sum(feed_ingredients.values())
        total_mix = sum(mix_ingredients.values())
        total_all = total_feed + total_mix

        overview_data = [
            ['Tổng nguyên liệu Feed (kg)', total_feed],
            ['Tổng nguyên liệu Mix (kg)', total_mix],
            ['Tổng tất cả nguyên liệu (kg)', total_all],
            ['Tỷ lệ Feed (%)', (total_feed / total_all * 100) if total_all > 0 else 0],
            ['Tỷ lệ Mix (%)', (total_mix / total_all * 100) if total_all > 0 else 0],
            ['Số loại nguyên liệu Feed', len(feed_ingredients)],
            ['Số loại nguyên liệu Mix', len(mix_ingredients)],
            ['Số loại nguyên liệu chung', len(all_ingredients)]
        ]

        overview_df = pd.DataFrame(overview_data, columns=['Chỉ Số', 'Giá Trị'])
        current_row = self.format_data_table(ws, overview_df, current_row, table_title="TỔNG QUAN SO SÁNH")

        return ws

    def create_production_summary_worksheet(self, wb: Workbook, report_date: str, report_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet tổng hợp sản xuất"""
        ws = wb.create_sheet(title="Tổng Hợp Sản Xuất")

        current_row = self.format_worksheet_header(ws, "TỔNG HỢP SẢN XUẤT FEED VÀ MIX", report_date)

        # Lấy dữ liệu
        raw_data = report_data.get('raw_data', {})
        efficiency = report_data.get('efficiency_metrics', {})

        # Tổng hợp sản xuất
        production_data = [
            ['Loại Sản Phẩm', 'Khối Lượng (kg)', 'Tỷ Lệ (%)', 'Số Loại Nguyên Liệu'],
            ['Feed (Cám)', efficiency.get('feed_total', 0), efficiency.get('feed_percentage', 0), len(raw_data.get('feed_ingredients', {}))],
            ['Mix', efficiency.get('mix_total', 0), efficiency.get('mix_percentage', 0), len(raw_data.get('mix_ingredients', {}))],
            ['Tổng Cộng', efficiency.get('total_consumption', 0), 100.0, len(set(raw_data.get('feed_ingredients', {}).keys()) | set(raw_data.get('mix_ingredients', {}).keys()))]
        ]

        production_df = pd.DataFrame(production_data[1:], columns=production_data[0])
        current_row = self.format_data_table(ws, production_df, current_row, table_title="TỔNG HỢP SẢN XUẤT")

        # So sánh sản xuất vs tiêu thụ
        comparison_data = [
            ['Chỉ Số', 'Feed (kg)', 'Mix (kg)', 'Tổng (kg)'],
            ['Tổng Sản Xuất', efficiency.get('feed_total', 0), efficiency.get('mix_total', 0), efficiency.get('total_consumption', 0)],
            ['Tiêu Thụ Theo Trại', efficiency.get('feed_usage_total', 0), efficiency.get('mix_usage_total', 0), efficiency.get('feed_usage_total', 0) + efficiency.get('mix_usage_total', 0)],
            ['Chênh Lệch', efficiency.get('feed_total', 0) - efficiency.get('feed_usage_total', 0), efficiency.get('mix_total', 0) - efficiency.get('mix_usage_total', 0), efficiency.get('total_consumption', 0) - (efficiency.get('feed_usage_total', 0) + efficiency.get('mix_usage_total', 0))],
            ['Hiệu Suất (%)', efficiency.get('feed_usage_percentage', 0), efficiency.get('mix_usage_percentage', 0), ((efficiency.get('feed_usage_total', 0) + efficiency.get('mix_usage_total', 0)) / efficiency.get('total_consumption', 1) * 100) if efficiency.get('total_consumption', 0) > 0 else 0]
        ]

        comparison_df = pd.DataFrame(comparison_data[1:], columns=comparison_data[0])
        current_row = self.format_data_table(ws, comparison_df, current_row, table_title="SO SÁNH SẢN XUẤT VÀ TIÊU THỤ")

        # Thông tin bổ sung
        additional_info = [
            ['Thông Tin', 'Giá Trị'],
            ['Tổng số lô sản xuất', raw_data.get('total_batches', 0)],
            ['Công thức mặc định', raw_data.get('default_formula', 'N/A')],
            ['Điểm hiệu quả', raw_data.get('efficiency_score', 0)],
            ['Điểm chất lượng', raw_data.get('quality_score', 0)],
            ['Tình trạng thời tiết', raw_data.get('weather', {}).get('description', 'N/A')],
            ['Tác động thời tiết', raw_data.get('weather', {}).get('impact', 1.0)]
        ]

        additional_df = pd.DataFrame(additional_info[1:], columns=additional_info[0])
        current_row = self.format_data_table(ws, additional_df, current_row, table_title="THÔNG TIN BỔ SUNG")

        return ws

    def _get_export_file_path(self, report_date: str, filename: str = None) -> Path:
        """Lấy đường dẫn file xuất"""
        if not filename:
            display_date = self._format_display_date(report_date).replace('/', '')
            filename = f"BaoCao_TieuThu_Cam_{report_date}.xlsx"

        # Đảm bảo có đuôi .xlsx
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        # Sử dụng thư mục tùy chỉnh nếu có
        export_folder = self.preferences.get_export_folder_path()
        if export_folder and Path(export_folder).exists():
            export_dir = Path(export_folder)
        else:
            # Use persistent path manager for consistent paths
            from src.utils.persistent_paths import persistent_path_manager
            export_dir = persistent_path_manager.exports_path
            export_dir.mkdir(parents=True, exist_ok=True)

        return export_dir / filename

    def export_daily_feed_report(self, report_date: str, filename: str = None,
                                include_shift_analysis: bool = True,
                                include_area_analysis: bool = True,
                                include_mix_analysis: bool = True,
                                include_ingredients_comparison: bool = True) -> Tuple[bool, str]:
        """Xuất báo cáo tiêu thụ cám hàng ngày ra Excel"""
        try:
            print(f"📊 [Daily Feed Export] Starting export for {report_date}...")

            # Lấy dữ liệu báo cáo từ cache/calculator
            report_data = self.calculator.calculate_daily_report(report_date)
            if not report_data:
                return False, f"Không thể tải dữ liệu báo cáo cho ngày {report_date}"

            # Tạo workbook
            wb = self.create_workbook(report_date)

            # Tạo các worksheet
            print("📋 [Daily Feed Export] Creating worksheets...")

            # 1. Tổng quan (sheet đầu tiên)
            self.create_summary_worksheet(wb, report_date, report_data)

            # 2. Tổng hợp sản xuất
            self.create_production_summary_worksheet(wb, report_date, report_data)

            # 3. Chi tiết tiêu thụ cám theo trại
            self.create_feed_consumption_worksheet(wb, report_date, report_data)

            # 4. Phân tích theo ca (tùy chọn)
            if include_shift_analysis:
                self.create_shift_analysis_worksheet(wb, report_date, report_data)

            # 5. Phân tích theo khu vực (tùy chọn)
            if include_area_analysis:
                self.create_area_analysis_worksheet(wb, report_date, report_data)

            # 6. Tiêu thụ mix (tùy chọn)
            if include_mix_analysis:
                self.create_mix_consumption_worksheet(wb, report_date, report_data)

            # 7. So sánh nguyên liệu (tùy chọn)
            if include_ingredients_comparison:
                self.create_ingredients_comparison_worksheet(wb, report_date, report_data)

            # Lưu file
            file_path = self._get_export_file_path(report_date, filename)
            wb.save(file_path)

            print(f"✅ [Daily Feed Export] Export completed: {file_path}")
            return True, f"Báo cáo đã được xuất thành công: {file_path}"

        except Exception as e:
            error_msg = f"Lỗi khi xuất báo cáo tiêu thụ cám: {str(e)}"
            print(f"❌ [Daily Feed Export] {error_msg}")
            import traceback
            traceback.print_exc()
            return False, error_msg

# Global instance
daily_feed_excel_exporter = DailyFeedExcelExporter()

# Convenience functions
def export_daily_feed_to_excel(report_date: str, filename: str = None,
                              include_shift_analysis: bool = True,
                              include_area_analysis: bool = True,
                              include_mix_analysis: bool = True,
                              include_ingredients_comparison: bool = True) -> Tuple[bool, str]:
    """Xuất báo cáo tiêu thụ cám hàng ngày ra Excel"""
    return daily_feed_excel_exporter.export_daily_feed_report(
        report_date, filename, include_shift_analysis, include_area_analysis,
        include_mix_analysis, include_ingredients_comparison
    )

def get_daily_feed_export_path(report_date: str, filename: str = None) -> str:
    """Lấy đường dẫn file xuất báo cáo tiêu thụ cám"""
    return str(daily_feed_excel_exporter._get_export_file_path(report_date, filename))

