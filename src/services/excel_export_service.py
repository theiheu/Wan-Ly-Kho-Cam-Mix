#!/usr/bin/env python3
"""
Excel Export Service - Dịch vụ xuất Excel với định dạng chuyên nghiệp
Hỗ trợ xuất nhiều worksheet, định dạng tiếng Việt, và tùy chỉnh giao diện
"""

import os
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

class ExcelExportService:
    """Dịch vụ xuất Excel với định dạng chuyên nghiệp"""

    def __init__(self):
        """Khởi tạo dịch vụ xuất Excel"""
        self.data_dir = Path("src/data")
        self.exports_dir = self.data_dir / "exports"

        # Đảm bảo thư mục xuất tồn tại
        self.exports_dir.mkdir(parents=True, exist_ok=True)

        # Định nghĩa các style cho Excel
        self.setup_styles()

    def setup_styles(self):
        """Thiết lập các style cho Excel"""
        # Font styles
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.subheader_font = Font(name='Arial', size=11, bold=True, color='000000')
        self.normal_font = Font(name='Arial', size=10, color='000000')
        self.title_font = Font(name='Arial', size=14, bold=True, color='000000')

        # Fill styles
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        self.alternate_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        # Border styles
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Alignment styles
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')

    def create_workbook(self, title: str = "Báo Cáo Kho Hàng") -> Workbook:
        """Tạo workbook mới với thiết lập cơ bản"""
        wb = Workbook()

        # Xóa worksheet mặc định
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        return wb

    def format_worksheet_header(self, ws: Worksheet, title: str, start_row: int = 1):
        """Định dạng header cho worksheet"""
        # Thêm tiêu đề
        ws.cell(row=start_row, column=1, value=title)
        title_cell = ws.cell(row=start_row, column=1)
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment

        # Thêm thời gian tạo
        ws.cell(row=start_row + 1, column=1, value=f"Tạo lúc: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        time_cell = ws.cell(row=start_row + 1, column=1)
        time_cell.font = self.normal_font
        time_cell.alignment = self.left_alignment

        return start_row + 3  # Trả về hàng tiếp theo để bắt đầu dữ liệu

    def format_data_table(self, ws: Worksheet, df: pd.DataFrame, start_row: int, start_col: int = 1):
        """Định dạng bảng dữ liệu với header và border"""
        if df.empty:
            return start_row

        # Thêm dữ liệu vào worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.thin_border

                # Định dạng header
                if r_idx == start_row:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
                else:
                    cell.font = self.normal_font
                    # Định dạng số
                    if isinstance(value, (int, float)) and value != 0:
                        cell.number_format = '#,##0.00'
                        cell.alignment = self.right_alignment
                    else:
                        cell.alignment = self.left_alignment

                    # Màu xen kẽ cho các hàng
                    if (r_idx - start_row) % 2 == 0:
                        cell.fill = self.alternate_fill

        # Tự động điều chỉnh độ rộng cột
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Giới hạn độ rộng tối đa
            ws.column_dimensions[column_letter].width = adjusted_width

        return start_row + len(df) + 1

    def create_inventory_worksheet(self, wb: Workbook, inventory_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet tồn kho"""
        ws = wb.create_sheet(title="Tồn Kho")

        current_row = self.format_worksheet_header(ws, "BÁO CÁO TỒN KHO")

        # Tồn kho cám
        if inventory_data.get('feed_inventory'):
            ws.cell(row=current_row, column=1, value="KHO CÁM")
            ws.cell(row=current_row, column=1).font = self.subheader_font
            current_row += 1

            feed_df = pd.DataFrame([
                {'Tên nguyên liệu': name, 'Số lượng': quantity, 'Đơn vị': 'kg'}
                for name, quantity in inventory_data['feed_inventory'].items()
            ])
            current_row = self.format_data_table(ws, feed_df, current_row) + 1

        # Tồn kho mix
        if inventory_data.get('mix_inventory'):
            ws.cell(row=current_row, column=1, value="KHO MIX")
            ws.cell(row=current_row, column=1).font = self.subheader_font
            current_row += 1

            mix_df = pd.DataFrame([
                {'Tên nguyên liệu': name, 'Số lượng': quantity, 'Đơn vị': 'kg'}
                for name, quantity in inventory_data['mix_inventory'].items()
            ])
            current_row = self.format_data_table(ws, mix_df, current_row) + 1

        # Tồn kho chung
        if inventory_data.get('general_inventory'):
            ws.cell(row=current_row, column=1, value="KHO CHUNG")
            ws.cell(row=current_row, column=1).font = self.subheader_font
            current_row += 1

            general_df = pd.DataFrame([
                {'Tên nguyên liệu': name, 'Số lượng': quantity, 'Đơn vị': 'kg'}
                for name, quantity in inventory_data['general_inventory'].items()
            ])
            current_row = self.format_data_table(ws, general_df, current_row) + 1

        # Cảnh báo tồn kho thấp
        if inventory_data.get('low_stock_items'):
            ws.cell(row=current_row, column=1, value="CẢNH BÁO TỒN KHO THẤP")
            ws.cell(row=current_row, column=1).font = self.subheader_font
            current_row += 1

            low_stock_df = pd.DataFrame([
                {'Tên nguyên liệu': item['name'], 'Số lượng còn lại': item['quantity']}
                for item in inventory_data['low_stock_items']
            ])
            current_row = self.format_data_table(ws, low_stock_df, current_row)

        return ws

    def create_employee_worksheet(self, wb: Workbook, employee_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet nhân viên"""
        ws = wb.create_sheet(title="Nhân Viên")

        current_row = self.format_worksheet_header(ws, "BÁO CÁO NHÂN VIÊN")

        # Danh sách nhân viên
        if employee_data.get('employees'):
            employees_df = pd.DataFrame(employee_data['employees'])
            if 'created_date' in employees_df.columns:
                employees_df['created_date'] = pd.to_datetime(employees_df['created_date']).dt.strftime('%d/%m/%Y')
            current_row = self.format_data_table(ws, employees_df, current_row) + 2

        # Thống kê theo vị trí
        if employee_data.get('positions'):
            ws.cell(row=current_row, column=1, value="THỐNG KÊ THEO VỊ TRÍ")
            ws.cell(row=current_row, column=1).font = self.subheader_font
            current_row += 1

            positions_df = pd.DataFrame([
                {'Vị trí': position, 'Số lượng': count}
                for position, count in employee_data['positions'].items()
            ])
            current_row = self.format_data_table(ws, positions_df, current_row)

        return ws

    def create_production_worksheet(self, wb: Workbook, production_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet sản xuất"""
        ws = wb.create_sheet(title="Sản Xuất")

        current_row = self.format_worksheet_header(ws, "BÁO CÁO SẢN XUẤT")

        # Thông tin tổng quan
        summary_data = [
            ['Tổng số báo cáo', production_data.get('total_reports', 0)],
            ['Tổng cám sử dụng (kg)', production_data.get('total_feed_usage', 0)],
            ['Tổng mix sử dụng (kg)', production_data.get('total_mix_usage', 0)]
        ]

        if production_data.get('daily_averages'):
            summary_data.extend([
                ['Trung bình cám/ngày (kg)', production_data['daily_averages'].get('feed_per_day', 0)],
                ['Trung bình mix/ngày (kg)', production_data['daily_averages'].get('mix_per_day', 0)]
            ])

        summary_df = pd.DataFrame(summary_data, columns=['Chỉ số', 'Giá trị'])
        current_row = self.format_data_table(ws, summary_df, current_row) + 2

        # Sử dụng cám theo khu
        if production_data.get('feed_usage_by_area'):
            ws.cell(row=current_row, column=1, value="SỬ DỤNG CÁM THEO KHU")
            ws.cell(row=current_row, column=1).font = self.subheader_font
            current_row += 1

            feed_area_df = pd.DataFrame([
                {'Khu vực': area, 'Tổng sử dụng (kg)': usage}
                for area, usage in production_data['feed_usage_by_area'].items()
            ])
            current_row = self.format_data_table(ws, feed_area_df, current_row) + 2

        # Sử dụng mix theo khu
        if production_data.get('mix_usage_by_area'):
            ws.cell(row=current_row, column=1, value="SỬ DỤNG MIX THEO KHU")
            ws.cell(row=current_row, column=1).font = self.subheader_font
            current_row += 1

            mix_area_df = pd.DataFrame([
                {'Khu vực': area, 'Tổng sử dụng (kg)': usage}
                for area, usage in production_data['mix_usage_by_area'].items()
            ])
            current_row = self.format_data_table(ws, mix_area_df, current_row)

        return ws

    def create_bonus_worksheet(self, wb: Workbook, bonus_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet thưởng"""
        ws = wb.create_sheet(title="Thưởng")

        current_row = self.format_worksheet_header(ws, "BÁO CÁO THƯỞNG")

        # Thông tin tổng quan
        summary_data = [
            ['Tổng số tháng', bonus_data.get('total_months', 0)]
        ]
        summary_df = pd.DataFrame(summary_data, columns=['Chỉ số', 'Giá trị'])
        current_row = self.format_data_table(ws, summary_df, current_row) + 2

        # Thưởng theo nguyên liệu
        if bonus_data.get('ingredient_totals'):
            ws.cell(row=current_row, column=1, value="TỔNG THƯỞNG THEO NGUYÊN LIỆU")
            ws.cell(row=current_row, column=1).font = self.subheader_font
            current_row += 1

            ingredient_df = pd.DataFrame([
                {'Nguyên liệu': ingredient, 'Tổng thưởng (VNĐ)': total}
                for ingredient, total in bonus_data['ingredient_totals'].items()
            ])
            current_row = self.format_data_table(ws, ingredient_df, current_row) + 2

        # Thưởng theo tháng
        if bonus_data.get('monthly_totals'):
            ws.cell(row=current_row, column=1, value="TỔNG THƯỞNG THEO THÁNG")
            ws.cell(row=current_row, column=1).font = self.subheader_font
            current_row += 1

            monthly_df = pd.DataFrame([
                {'Tháng': month, 'Tổng thưởng (VNĐ)': total}
                for month, total in bonus_data['monthly_totals'].items()
            ])
            current_row = self.format_data_table(ws, monthly_df, current_row)

        return ws

    def create_formula_worksheet(self, wb: Workbook, formula_data: Dict[str, Any]) -> Worksheet:
        """Tạo worksheet công thức"""
        ws = wb.create_sheet(title="Công Thức")

        current_row = self.format_worksheet_header(ws, "BÁO CÁO CÔNG THỨC")

        # Thông tin tổng quan
        summary_data = [
            ['Tổng số công thức', formula_data.get('total_formulas', 0)],
            ['Số công thức cám', len(formula_data.get('feed_formulas', {}))],
            ['Số công thức mix', len(formula_data.get('mix_formulas', {}))]
        ]
        summary_df = pd.DataFrame(summary_data, columns=['Chỉ số', 'Giá trị'])
        current_row = self.format_data_table(ws, summary_df, current_row) + 2

        # Công thức cám
        if formula_data.get('feed_formulas'):
            ws.cell(row=current_row, column=1, value="CÔNG THỨC CÁM")
            ws.cell(row=current_row, column=1).font = self.subheader_font
            current_row += 1

            feed_formulas = []
            for formula_name, ingredients in formula_data['feed_formulas'].items():
                if isinstance(ingredients, dict):
                    for ingredient, amount in ingredients.items():
                        feed_formulas.append({
                            'Tên công thức': formula_name,
                            'Nguyên liệu': ingredient,
                            'Tỷ lệ (%)': amount
                        })

            if feed_formulas:
                feed_df = pd.DataFrame(feed_formulas)
                current_row = self.format_data_table(ws, feed_df, current_row) + 2

        # Công thức mix
        if formula_data.get('mix_formulas'):
            ws.cell(row=current_row, column=1, value="CÔNG THỨC MIX")
            ws.cell(row=current_row, column=1).font = self.subheader_font
            current_row += 1

            mix_formulas = []
            for formula_name, ingredients in formula_data['mix_formulas'].items():
                if isinstance(ingredients, dict):
                    for ingredient, amount in ingredients.items():
                        mix_formulas.append({
                            'Tên công thức': formula_name,
                            'Nguyên liệu': ingredient,
                            'Tỷ lệ (%)': amount
                        })

            if mix_formulas:
                mix_df = pd.DataFrame(mix_formulas)
                current_row = self.format_data_table(ws, mix_df, current_row)

        return ws

    def export_comprehensive_report(self, report_data: Dict[str, Any], filename: str = None, custom_export_dir: str = None) -> Tuple[bool, str]:
        """Xuất báo cáo toàn diện ra Excel"""
        try:
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"BaoCao_ToanDien_{timestamp}.xlsx"

            # Đảm bảo file có đuôi .xlsx
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'

            # Sử dụng thư mục tùy chỉnh nếu được cung cấp
            if custom_export_dir and Path(custom_export_dir).exists():
                export_directory = Path(custom_export_dir)
            else:
                export_directory = self.exports_dir

            file_path = export_directory / filename

            # Tạo workbook
            wb = self.create_workbook("Báo Cáo Toàn Diện")

            # Tạo các worksheet dựa trên dữ liệu có sẵn
            sections = report_data.get('sections', {})

            if 'inventory' in sections:
                self.create_inventory_worksheet(wb, sections['inventory'])

            if 'employees' in sections:
                self.create_employee_worksheet(wb, sections['employees'])

            if 'production' in sections:
                self.create_production_worksheet(wb, sections['production'])

            if 'bonuses' in sections:
                self.create_bonus_worksheet(wb, sections['bonuses'])

            if 'formulas' in sections:
                self.create_formula_worksheet(wb, sections['formulas'])

            # Tạo worksheet tổng quan
            self.create_summary_worksheet(wb, report_data)

            # Lưu file
            wb.save(file_path)

            return True, f"Báo cáo đã được xuất thành công: {file_path}"

        except Exception as e:
            return False, f"Lỗi khi xuất báo cáo: {str(e)}"

    def create_summary_worksheet(self, wb: Workbook, report_data: Dict[str, Any]):
        """Tạo worksheet tổng quan"""
        ws = wb.create_sheet(title="Tổng Quan", index=0)  # Đặt làm sheet đầu tiên

        current_row = self.format_worksheet_header(ws, "BÁO CÁO TỔNG QUAN HỆ THỐNG")

        # Thông tin báo cáo
        info_data = [
            ['Thời gian tạo', report_data.get('generated_at', '')],
            ['Khoảng thời gian', f"{report_data.get('date_range', {}).get('start', 'N/A')} - {report_data.get('date_range', {}).get('end', 'N/A')}"]
        ]

        sections = report_data.get('sections', {})

        # Thống kê từng phần
        if 'inventory' in sections:
            inv_data = sections['inventory']
            info_data.extend([
                ['Tổng số mặt hàng tồn kho', inv_data.get('total_items', 0)],
                ['Số mặt hàng sắp hết', len(inv_data.get('low_stock_items', []))],
                ['Số mặt hàng đã hết', len(inv_data.get('out_of_stock_items', []))]
            ])

        if 'employees' in sections:
            emp_data = sections['employees']
            info_data.append(['Tổng số nhân viên', emp_data.get('total_employees', 0)])

        if 'production' in sections:
            prod_data = sections['production']
            info_data.extend([
                ['Tổng số báo cáo sản xuất', prod_data.get('total_reports', 0)],
                ['Tổng cám sử dụng (kg)', prod_data.get('total_feed_usage', 0)],
                ['Tổng mix sử dụng (kg)', prod_data.get('total_mix_usage', 0)]
            ])

        if 'bonuses' in sections:
            bonus_data = sections['bonuses']
            info_data.append(['Số tháng có dữ liệu thưởng', bonus_data.get('total_months', 0)])

        if 'formulas' in sections:
            formula_data = sections['formulas']
            info_data.append(['Tổng số công thức', formula_data.get('total_formulas', 0)])

        summary_df = pd.DataFrame(info_data, columns=['Thông tin', 'Giá trị'])
        self.format_data_table(ws, summary_df, current_row)

        return ws
