#!/usr/bin/env python3
"""
Optimized Export Service - D·ªãch v·ª• xu·∫•t b√°o c√°o ƒë∆∞·ª£c t·ªëi ∆∞u h√≥a
C·∫£i thi·ªán hi·ªáu su·∫•t, ch·∫•t l∆∞·ª£ng Excel v√† tr·∫£i nghi·ªám ng∆∞·ªùi d√πng
"""

import os
import json
import pandas as pd
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from concurrent.futures import ThreadPoolExecutor
import threading

# Excel formatting imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image


class ExcelStyleManager:
    """Qu·∫£n l√Ω styles cho Excel v·ªõi hi·ªáu su·∫•t cao"""

    def __init__(self):
        self._styles_cache = {}
        self._create_predefined_styles()

    def _create_predefined_styles(self):
        """T·∫°o s·∫µn c√°c style th∆∞·ªùng d√πng"""
        # Header style
        self.header_style = NamedStyle(name="header_style")
        self.header_style.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_style.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.header_style.alignment = Alignment(horizontal='center', vertical='center')

        # Data styles
        self.data_style = NamedStyle(name="data_style")
        self.data_style.font = Font(name='Arial', size=10)
        self.data_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.data_style.alignment = Alignment(horizontal='left', vertical='center')

        # Number style
        self.number_style = NamedStyle(name="number_style")
        self.number_style.font = Font(name='Arial', size=10)
        self.number_style.number_format = '#,##0.00'
        self.number_style.alignment = Alignment(horizontal='right', vertical='center')

        # Status styles
        self.status_styles = {
            'ƒê·ªß': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            '√çt': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'S·∫Øp h·∫øt': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'H·∫øt h√†ng': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        }

    def get_status_fill(self, status: str) -> PatternFill:
        """L·∫•y fill color cho status"""
        return self.status_styles.get(status, PatternFill())


class OptimizedExportService:
    """D·ªãch v·ª• xu·∫•t b√°o c√°o ƒë∆∞·ª£c t·ªëi ∆∞u h√≥a"""

    def __init__(self, inventory_manager=None, formula_manager=None, threshold_manager=None, remaining_usage_calculator=None):
        """Kh·ªüi t·∫°o d·ªãch v·ª• v·ªõi data managers t·ª´ ·ª©ng d·ª•ng ch√≠nh"""
        # Use persistent path manager for consistent paths
        from src.utils.persistent_paths import persistent_path_manager

        self.data_dir = persistent_path_manager.data_path
        self.config_dir = persistent_path_manager.config_path
        self.exports_dir = persistent_path_manager.exports_path
        self.reports_dir = persistent_path_manager.reports_path
        self.daily_consumption_dir = self.data_dir / "daily_consumption"

        print(f"üîß OptimizedExportService initialized:")
        print(f"   üìÅ Data dir: {self.data_dir}")
        print(f"   üìÅ Config dir: {self.config_dir}")
        print(f"   üìÅ Exports dir: {self.exports_dir}")
        print(f"   üìÅ Reports dir: {self.reports_dir}")

        # Performance optimizations
        self._data_cache = {}
        self._cache_lock = threading.Lock()
        self._cache_timeout = 300  # 5 minutes

        # Style manager
        self.style_manager = ExcelStyleManager()

        # Ensure directories exist
        for directory in [self.exports_dir, self.daily_consumption_dir, self.reports_dir]:
            directory.mkdir(parents=True, exist_ok=True)

        # Data managers from main application
        self.inventory_manager = inventory_manager
        self.formula_manager = formula_manager
        self.threshold_manager = threshold_manager
        self.remaining_usage_calculator = remaining_usage_calculator

        # Initialize data managers if not provided (fallback to sample data)
        if not self.inventory_manager or not self.formula_manager:
            print("‚ö†Ô∏è Warning: Data managers not provided, falling back to sample data")
            self._ensure_sample_data()

        # Ensure daily consumption data exists
        self._ensure_daily_consumption_data()

    def get_real_time_inventory_summary(self) -> Dict:
        """L·∫•y t√≥m t·∫Øt t·ªìn kho real-time t·ª´ inventory manager"""
        try:
            if not self.inventory_manager:
                return {}

            summary = {
                'total_items': 0,
                'total_value': 0,
                'low_stock_items': 0,
                'out_of_stock_items': 0,
                'warehouses': {}
            }

            # Get data for each warehouse type
            for warehouse_type in ['feed', 'mix']:
                warehouse_data = self._get_real_inventory_data(warehouse_type)
                warehouse_summary = self._analyze_warehouse_data(warehouse_data, warehouse_type)
                summary['warehouses'][warehouse_type] = warehouse_summary

                # Add to totals
                summary['total_items'] += warehouse_summary.get('item_count', 0)
                summary['total_value'] += warehouse_summary.get('total_value', 0)
                summary['low_stock_items'] += warehouse_summary.get('low_stock_count', 0)
                summary['out_of_stock_items'] += warehouse_summary.get('out_of_stock_count', 0)

            return summary

        except Exception as e:
            print(f"Error getting real-time inventory summary: {e}")
            return {}

    def get_real_time_production_capacity(self) -> Dict:
        """L·∫•y kh·∫£ nƒÉng s·∫£n xu·∫•t real-time t·ª´ formula v√† inventory managers"""
        try:
            if not self.formula_manager or not self.inventory_manager:
                return {}

            capacity = {
                'feed_capacity': {},
                'mix_capacity': {},
                'bottleneck_ingredients': []
            }

            # Analyze feed production capacity
            feed_formula = self._get_real_formula_data("feed")
            feed_inventory = self._get_real_inventory_data("feed")
            capacity['feed_capacity'] = self._calculate_production_capacity(feed_formula, feed_inventory)

            # Analyze mix production capacity
            mix_formula = self._get_real_formula_data("mix")
            mix_inventory = self._get_real_inventory_data("mix")
            capacity['mix_capacity'] = self._calculate_production_capacity(mix_formula, mix_inventory)

            # Identify bottleneck ingredients
            capacity['bottleneck_ingredients'] = self._identify_bottlenecks(
                [capacity['feed_capacity'], capacity['mix_capacity']]
            )

            return capacity

        except Exception as e:
            print(f"Error getting real-time production capacity: {e}")
            return {}

    def get_real_time_usage_trends(self, days: int = 30) -> Dict:
        """L·∫•y xu h∆∞·ªõng s·ª≠ d·ª•ng real-time"""
        try:
            if not self.remaining_usage_calculator:
                return {}

            trends = {
                'consumption_trends': {},
                'prediction_accuracy': {},
                'recommended_orders': []
            }

            # Get consumption trends
            trends['consumption_trends'] = self.remaining_usage_calculator.get_consumption_trends(days)

            # Get prediction accuracy
            trends['prediction_accuracy'] = self.remaining_usage_calculator.get_prediction_accuracy()

            # Get recommended orders
            trends['recommended_orders'] = self.remaining_usage_calculator.get_recommended_orders()

            return trends

        except Exception as e:
            print(f"Error getting real-time usage trends: {e}")
            return {}

    def _analyze_warehouse_data(self, warehouse_data: Dict, warehouse_type: str) -> Dict:
        """Ph√¢n t√≠ch d·ªØ li·ªáu kho"""
        try:
            analysis = {
                'item_count': 0,
                'total_quantity': 0,
                'total_value': 0,
                'low_stock_count': 0,
                'out_of_stock_count': 0,
                'categories': {}
            }

            unit_price = 15000 if warehouse_type == 'feed' else 25000

            if isinstance(warehouse_data, dict):
                if 'items' in warehouse_data:
                    items = warehouse_data['items']
                elif 'inventory' in warehouse_data:
                    items = [{'name': k, 'quantity': v} for k, v in warehouse_data['inventory'].items()]
                else:
                    items = [{'name': k, 'quantity': v} for k, v in warehouse_data.items()]

                for item in items:
                    if isinstance(item, dict):
                        quantity = float(item.get('quantity', 0))
                        analysis['item_count'] += 1
                        analysis['total_quantity'] += quantity
                        analysis['total_value'] += quantity * unit_price

                        # Check stock status
                        if quantity == 0:
                            analysis['out_of_stock_count'] += 1
                        elif quantity < 100:  # Low stock threshold
                            analysis['low_stock_count'] += 1

            return analysis

        except Exception as e:
            print(f"Error analyzing warehouse data: {e}")
            return {}

    def _calculate_production_capacity(self, formula_data: Dict, inventory_data: Dict) -> Dict:
        """T√≠nh to√°n kh·∫£ nƒÉng s·∫£n xu·∫•t"""
        try:
            capacity = {
                'max_batches': 0,
                'limiting_ingredients': [],
                'ingredient_usage': {}
            }

            if not formula_data or not inventory_data:
                return capacity

            min_batches = float('inf')

            # Process formula data
            if isinstance(formula_data, dict):
                if 'ingredients' in formula_data:
                    ingredients = formula_data['ingredients']
                elif 'formula' in formula_data:
                    ingredients = [{'name': k, 'percentage': v} for k, v in formula_data['formula'].items()]
                else:
                    ingredients = [{'name': k, 'percentage': v} for k, v in formula_data.items()]

                for ingredient in ingredients:
                    if isinstance(ingredient, dict):
                        name = ingredient.get('name', '')
                        percentage = float(ingredient.get('percentage', 0))

                        if percentage > 0:
                            current_stock = self._get_ingredient_stock(name, inventory_data)
                            usage_per_batch = (percentage / 100) * 1000  # For 1000kg batch
                            possible_batches = int(current_stock / usage_per_batch) if usage_per_batch > 0 else 0

                            capacity['ingredient_usage'][name] = {
                                'stock': current_stock,
                                'usage_per_batch': usage_per_batch,
                                'possible_batches': possible_batches
                            }

                            if possible_batches < min_batches:
                                min_batches = possible_batches
                                capacity['limiting_ingredients'] = [name]
                            elif possible_batches == min_batches:
                                capacity['limiting_ingredients'].append(name)

            capacity['max_batches'] = min_batches if min_batches != float('inf') else 0
            return capacity

        except Exception as e:
            print(f"Error calculating production capacity: {e}")
            return {}

    def _identify_bottlenecks(self, capacity_data: List[Dict]) -> List[Dict]:
        """X√°c ƒë·ªãnh c√°c nguy√™n li·ªáu c·ªï chai"""
        try:
            bottlenecks = []

            for capacity in capacity_data:
                for ingredient, usage_data in capacity.get('ingredient_usage', {}).items():
                    possible_batches = usage_data.get('possible_batches', 0)
                    if possible_batches < 10:  # Less than 10 batches possible
                        bottlenecks.append({
                            'ingredient': ingredient,
                            'current_stock': usage_data.get('stock', 0),
                            'possible_batches': possible_batches,
                            'urgency': 'high' if possible_batches < 5 else 'medium'
                        })

            # Sort by urgency and possible batches
            bottlenecks.sort(key=lambda x: (x['urgency'] == 'high', -x['possible_batches']))

            return bottlenecks

        except Exception as e:
            print(f"Error identifying bottlenecks: {e}")
            return []

    def _get_real_inventory_data(self, warehouse_type: str = None) -> Dict:
        """L·∫•y d·ªØ li·ªáu t·ªìn kho th·ª±c t·ª´ inventory_manager"""
        try:
            if not self.inventory_manager:
                print("‚ö†Ô∏è No inventory_manager available, using sample data")
                return self._load_json_cached(self.config_dir / f"{warehouse_type}_inventory.json" if warehouse_type else self.config_dir / "feed_inventory.json")

            if warehouse_type:
                # Get specific warehouse inventory
                return self.inventory_manager.get_warehouse_inventory(warehouse_type)
            else:
                # Get all inventory
                return self.inventory_manager.get_inventory()

        except Exception as e:
            print(f"Error getting real inventory data: {e}")
            # Fallback to sample data
            return self._load_json_cached(self.config_dir / f"{warehouse_type}_inventory.json" if warehouse_type else self.config_dir / "feed_inventory.json")

    def _get_real_formula_data(self, formula_type: str) -> Dict:
        """L·∫•y d·ªØ li·ªáu c√¥ng th·ª©c th·ª±c t·ª´ formula_manager"""
        try:
            if not self.formula_manager:
                print("‚ö†Ô∏è No formula_manager available, using sample data")
                return self._load_json_cached(self.config_dir / f"{formula_type}_formula.json")

            if formula_type == "feed":
                return self.formula_manager.get_feed_formula()
            elif formula_type == "mix":
                return self.formula_manager.get_mix_formula()
            else:
                return {}

        except Exception as e:
            print(f"Error getting real formula data: {e}")
            # Fallback to sample data
            return self._load_json_cached(self.config_dir / f"{formula_type}_formula.json")

    def _get_real_packaging_info(self) -> Dict:
        """L·∫•y th√¥ng tin ƒë√≥ng g√≥i th·ª±c t·ª´ inventory_manager"""
        try:
            if not self.inventory_manager:
                return {}

            return self.inventory_manager.get_packaging_info()

        except Exception as e:
            print(f"Error getting real packaging info: {e}")
            return {}

    def _get_real_usage_analysis(self) -> Dict:
        """L·∫•y ph√¢n t√≠ch s·ª≠ d·ª•ng th·ª±c t·ª´ remaining_usage_calculator"""
        try:
            if not self.remaining_usage_calculator:
                return {}

            # Get usage analysis for both feed and mix
            return self.remaining_usage_calculator.analyze_usage_patterns()

        except Exception as e:
            print(f"Error getting real usage analysis: {e}")
            return {}

    def _process_inventory_data(self, export_data: List[Dict], inventory_data: Dict, warehouse_type: str, unit_price: float):
        """X·ª≠ l√Ω d·ªØ li·ªáu t·ªìn kho t·ª´ inventory manager"""
        try:
            if isinstance(inventory_data, dict):
                # Handle different data structures from inventory manager
                if 'items' in inventory_data:
                    # Structure: {'items': [{'name': ..., 'quantity': ..., 'unit': ...}, ...]}
                    for item in inventory_data.get('items', []):
                        self._add_inventory_item(export_data, item, warehouse_type, unit_price)
                elif 'inventory' in inventory_data:
                    # Structure: {'inventory': {'item_name': quantity, ...}}
                    for item_name, quantity in inventory_data.get('inventory', {}).items():
                        item_data = {'name': item_name, 'quantity': quantity}
                        self._add_inventory_item(export_data, item_data, warehouse_type, unit_price)
                else:
                    # Simple structure: {'item_name': quantity, ...}
                    for item_name, quantity in inventory_data.items():
                        item_data = {'name': item_name, 'quantity': quantity}
                        self._add_inventory_item(export_data, item_data, warehouse_type, unit_price)
            elif isinstance(inventory_data, list):
                # Handle list of items
                for item in inventory_data:
                    self._add_inventory_item(export_data, item, warehouse_type, unit_price)

        except Exception as e:
            print(f"Error processing inventory data: {e}")

    def _add_inventory_item(self, export_data: List[Dict], item_data: Dict, warehouse_type: str, unit_price: float):
        """Th√™m m·ªôt item t·ªìn kho v√†o export data"""
        try:
            # Extract item information with fallbacks
            item_name = item_data.get('name', item_data.get('item_name', 'Unknown'))
            quantity = float(item_data.get('quantity', item_data.get('amount', 0)))
            unit = item_data.get('unit', 'kg')

            # Get additional info if available
            supplier = item_data.get('supplier', item_data.get('nha_cung_cap', ''))
            batch_number = item_data.get('batch_number', item_data.get('so_lo', ''))
            expiry_date = item_data.get('expiry_date', item_data.get('han_su_dung', ''))

            # Get stock status
            status = self._get_stock_status_from_item(item_data, quantity)

            # Calculate estimated value
            estimated_value = quantity * unit_price

            export_data.append({
                'Lo·∫°i Kho': warehouse_type,
                'T√™n Nguy√™n Li·ªáu': item_name,
                'S·ªë L∆∞·ª£ng': quantity,
                'ƒê∆°n V·ªã': unit,
                'Tr·∫°ng Th√°i': status,
                'Nh√† Cung C·∫•p': supplier,
                'S·ªë L√¥': batch_number,
                'H·∫°n S·ª≠ D·ª•ng': expiry_date,
                'Gi√° Tr·ªã ∆Ø·ªõc T√≠nh': estimated_value,
                'Ng√†y C·∫≠p Nh·∫≠t': datetime.now().strftime("%d/%m/%Y %H:%M")
            })

        except Exception as e:
            print(f"Error adding inventory item: {e}")

    def _get_stock_status_from_item(self, item_data: Dict, quantity: float) -> str:
        """L·∫•y tr·∫°ng th√°i t·ªìn kho t·ª´ item data ho·∫∑c t√≠nh to√°n"""
        try:
            # Check if status is already provided
            if 'status' in item_data:
                return item_data['status']
            if 'trang_thai' in item_data:
                return item_data['trang_thai']

            # Use threshold manager if available
            if self.threshold_manager:
                item_name = item_data.get('name', item_data.get('item_name', ''))
                return self.threshold_manager.get_stock_status(item_name, quantity)

            # Fallback to simple quantity-based status
            return self._get_stock_status(quantity)

        except Exception as e:
            print(f"Error getting stock status: {e}")
            return self._get_stock_status(quantity)

    def _process_formula_data(self, export_data: List[Dict], formula_data: Dict, inventory_data: Dict, formula_type: str):
        """X·ª≠ l√Ω d·ªØ li·ªáu c√¥ng th·ª©c t·ª´ formula manager"""
        try:
            if isinstance(formula_data, dict):
                # Handle different formula data structures
                if 'ingredients' in formula_data:
                    # Structure: {'ingredients': [{'name': ..., 'percentage': ...}, ...]}
                    for ingredient in formula_data.get('ingredients', []):
                        self._add_formula_ingredient(export_data, ingredient, inventory_data, formula_type)
                elif 'formula' in formula_data:
                    # Structure: {'formula': {'ingredient_name': percentage, ...}}
                    for ingredient_name, percentage in formula_data.get('formula', {}).items():
                        ingredient_data = {'name': ingredient_name, 'percentage': percentage}
                        self._add_formula_ingredient(export_data, ingredient_data, inventory_data, formula_type)
                else:
                    # Simple structure: {'ingredient_name': percentage, ...}
                    for ingredient_name, percentage in formula_data.items():
                        ingredient_data = {'name': ingredient_name, 'percentage': percentage}
                        self._add_formula_ingredient(export_data, ingredient_data, inventory_data, formula_type)
            elif isinstance(formula_data, list):
                # Handle list of ingredients
                for ingredient in formula_data:
                    self._add_formula_ingredient(export_data, ingredient, inventory_data, formula_type)

        except Exception as e:
            print(f"Error processing formula data: {e}")

    def _add_formula_ingredient(self, export_data: List[Dict], ingredient_data: Dict, inventory_data: Dict, formula_type: str):
        """Th√™m m·ªôt ingredient c√¥ng th·ª©c v√†o export data"""
        try:
            # Extract ingredient information
            ingredient_name = ingredient_data.get('name', ingredient_data.get('ingredient_name', 'Unknown'))
            percentage = float(ingredient_data.get('percentage', ingredient_data.get('ratio', 0)))

            # Get current stock from inventory
            current_stock = self._get_ingredient_stock(ingredient_name, inventory_data)

            # Calculate usage and production capacity
            usage_per_batch = (percentage / 100) * 1000  # For 1000kg batch
            possible_batches = int(current_stock / usage_per_batch) if usage_per_batch > 0 else 0

            # Get additional info if available
            nutritional_value = ingredient_data.get('nutritional_value', ingredient_data.get('gia_tri_dinh_duong', ''))
            function = ingredient_data.get('function', ingredient_data.get('chuc_nang', ''))

            # Determine ingredient category
            category = self._categorize_ingredient(ingredient_name, percentage)

            export_data.append({
                'Lo·∫°i C√¥ng Th·ª©c': formula_type,
                'Nguy√™n Li·ªáu': ingredient_name,
                'T·ª∑ L·ªá (%)': percentage,
                'Kh·ªëi L∆∞·ª£ng/1000kg': round(usage_per_batch, 2),
                'T·ªìn Kho Hi·ªán T·∫°i (kg)': current_stock,
                'S·ªë L√¥ C√≥ Th·ªÉ S·∫£n Xu·∫•t': possible_batches,
                'Tr·∫°ng Th√°i Nguy√™n Li·ªáu': self._get_stock_status(current_stock),
                'Gi√° Tr·ªã Dinh D∆∞·ª°ng': nutritional_value,
                'Ch·ª©c NƒÉng': function,
                'Ghi Ch√∫': category
            })

        except Exception as e:
            print(f"Error adding formula ingredient: {e}")

    def _get_ingredient_stock(self, ingredient_name: str, inventory_data: Dict) -> float:
        """L·∫•y s·ªë l∆∞·ª£ng t·ªìn kho c·ªßa nguy√™n li·ªáu"""
        try:
            if isinstance(inventory_data, dict):
                # Handle different inventory structures
                if 'items' in inventory_data:
                    for item in inventory_data.get('items', []):
                        if item.get('name', '') == ingredient_name:
                            return float(item.get('quantity', 0))
                elif 'inventory' in inventory_data:
                    return float(inventory_data.get('inventory', {}).get(ingredient_name, 0))
                else:
                    return float(inventory_data.get(ingredient_name, 0))
            return 0.0
        except Exception as e:
            print(f"Error getting ingredient stock: {e}")
            return 0.0

    def _categorize_ingredient(self, ingredient_name: str, percentage: float) -> str:
        """Ph√¢n lo·∫°i nguy√™n li·ªáu d·ª±a tr√™n t√™n v√† t·ª∑ l·ªá"""
        try:
            ingredient_lower = ingredient_name.lower()

            if percentage >= 20:
                return 'Nguy√™n li·ªáu ch√≠nh'
            elif 'vitamin' in ingredient_lower or 'khoang' in ingredient_lower:
                return 'Vitamin/Kho√°ng ch·∫•t'
            elif 'enzyme' in ingredient_lower or 'men' in ingredient_lower:
                return 'Enzyme/Men vi sinh'
            elif percentage < 1:
                return 'Vi l∆∞·ª£ng'
            else:
                return 'Nguy√™n li·ªáu ph·ª•'

        except Exception as e:
            print(f"Error categorizing ingredient: {e}")
            return 'Nguy√™n li·ªáu ph·ª•'

    def _load_json_cached(self, file_path: Path) -> Dict:
        """T·∫£i JSON v·ªõi caching ƒë·ªÉ tƒÉng hi·ªáu su·∫•t"""
        cache_key = str(file_path)
        current_time = time.time()

        with self._cache_lock:
            # Check cache
            if cache_key in self._data_cache:
                cached_data, cache_time = self._data_cache[cache_key]
                if current_time - cache_time < self._cache_timeout:
                    return cached_data

            # Load from file
            try:
                if file_path.exists():
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        self._data_cache[cache_key] = (data, current_time)
                        return data
            except Exception as e:
                print(f"Error loading {file_path}: {e}")

            return {}

    def _clear_cache(self):
        """X√≥a cache khi c·∫ßn"""
        with self._cache_lock:
            self._data_cache.clear()

    def _ensure_sample_data(self):
        """T·∫°o d·ªØ li·ªáu m·∫´u v·ªõi nhi·ªÅu items h∆°n ƒë·ªÉ test performance"""
        self.config_dir.mkdir(parents=True, exist_ok=True)

        # Enhanced sample data for performance testing
        feed_inventory_file = self.config_dir / "feed_inventory.json"
        if not feed_inventory_file.exists() or self._is_file_empty(feed_inventory_file):
            sample_feed = {
                "B·∫Øp nghi·ªÅn": 1500.0, "C√°m g·∫°o": 800.0, "ƒê·∫≠u n√†nh": 1200.0,
                "D·∫ßu ƒë·∫≠u n√†nh": 150.0, "C√°m l√∫a m√¨": 600.0, "Kh√¥ d·ª´a": 300.0,
                "B·ªôt c√°": 250.0, "T·∫•m g·∫°o": 400.0, "R·ªâ m·∫≠t": 180.0,
                "B·ªôt x∆∞∆°ng": 120.0, "C√°m m√¨": 350.0, "Ng√¥ v√†ng": 900.0,
                "ƒê·∫≠u xanh": 200.0, "L√∫a m√¨": 750.0, "Y·∫øn m·∫°ch": 180.0
            }
            self._save_json(feed_inventory_file, sample_feed)

        mix_inventory_file = self.config_dir / "mix_inventory.json"
        if not mix_inventory_file.exists() or self._is_file_empty(mix_inventory_file):
            sample_mix = {
                "Lysine": 75.0, "Methionine": 45.0, "Choline": 60.0,
                "ƒê√° v√¥i": 300.0, "Vitamin premix": 25.0, "Mineral premix": 35.0,
                "Threonine": 20.0, "Tryptophan": 15.0, "Valine": 18.0,
                "Phytase": 12.0, "Xylanase": 8.0, "Protease": 10.0,
                "Antioxidant": 5.0, "Acidifier": 22.0, "Probiotic": 30.0
            }
            self._save_json(mix_inventory_file, sample_mix)

        # Enhanced formulas
        feed_formula_file = self.config_dir / "feed_formula.json"
        if not feed_formula_file.exists():
            sample_feed_formula = {
                "B·∫Øp nghi·ªÅn": 35.0, "C√°m g·∫°o": 15.0, "ƒê·∫≠u n√†nh": 20.0,
                "D·∫ßu ƒë·∫≠u n√†nh": 3.0, "C√°m l√∫a m√¨": 12.0, "Kh√¥ d·ª´a": 5.0,
                "B·ªôt c√°": 4.0, "T·∫•m g·∫°o": 3.0, "Nguy√™n li·ªáu t·ªï h·ª£p": 3.0
            }
            self._save_json(feed_formula_file, sample_feed_formula)

        mix_formula_file = self.config_dir / "mix_formula.json"
        if not mix_formula_file.exists():
            sample_mix_formula = {
                "Lysine": 25.0, "Methionine": 15.0, "Choline": 12.0,
                "ƒê√° v√¥i": 20.0, "Vitamin premix": 10.0, "Mineral premix": 8.0,
                "Threonine": 5.0, "Tryptophan": 3.0, "Phytase": 2.0
            }
            self._save_json(mix_formula_file, sample_mix_formula)

    def _is_file_empty(self, file_path: Path) -> bool:
        """Ki·ªÉm tra file c√≥ r·ªóng kh√¥ng"""
        try:
            if file_path.stat().st_size == 0:
                return True
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return not data or len(data) == 0
        except:
            return True

    def _save_json(self, file_path: Path, data: Dict):
        """L∆∞u d·ªØ li·ªáu JSON"""
        try:
            file_path.parent.mkdir(parents=True, exist_ok=True)
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Error saving {file_path}: {e}")

    def _get_stock_status(self, quantity: float) -> str:
        """X√°c ƒë·ªãnh tr·∫°ng th√°i t·ªìn kho v·ªõi logic c·∫£i ti·∫øn"""
        if quantity <= 0:
            return "H·∫øt h√†ng"
        elif quantity < 50:
            return "S·∫Øp h·∫øt"
        elif quantity < 200:
            return "√çt"
        else:
            return "ƒê·ªß"

    def _generate_filename(self, report_type: str, options: Dict = None) -> str:
        """T·∫°o t√™n file v·ªõi th√¥ng tin chi ti·∫øt h∆°n"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if options:
            suffix_parts = []
            if options.get('include_feed', False):
                suffix_parts.append('cam')
            if options.get('include_mix', False):
                suffix_parts.append('mix')

            if suffix_parts:
                suffix = '_' + '_'.join(suffix_parts)
            else:
                suffix = ''
        else:
            suffix = ''

        return f"{report_type}_{timestamp}{suffix}.xlsx"

    def _apply_advanced_formatting(self, worksheet, data_range: str, report_type: str):
        """√Åp d·ª•ng formatting n√¢ng cao cho worksheet"""
        # Apply header style
        for cell in worksheet[1]:
            cell.style = self.style_manager.header_style

        # Apply data styles with conditional formatting
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.column == worksheet.max_column and report_type == "inventory":
                    # Status column - apply conditional formatting
                    status = str(cell.value)
                    cell.fill = self.style_manager.get_status_fill(status)

                # Apply appropriate style based on data type
                if isinstance(cell.value, (int, float)):
                    cell.style = self.style_manager.number_style
                else:
                    cell.style = self.style_manager.data_style

        # Auto-adjust column widths (optimized)
        self._optimize_column_widths(worksheet)

        # Add alternating row colors
        self._add_alternating_rows(worksheet)

    def _optimize_column_widths(self, worksheet):
        """T·ªëi ∆∞u h√≥a ƒë·ªô r·ªông c·ªôt v·ªõi algorithm hi·ªáu qu·∫£ h∆°n"""
        column_widths = {}

        # Calculate optimal widths in one pass
        for row in worksheet.iter_rows():
            for cell in row:
                column_letter = cell.column_letter
                if column_letter not in column_widths:
                    column_widths[column_letter] = 0

                try:
                    cell_length = len(str(cell.value))
                    if cell_length > column_widths[column_letter]:
                        column_widths[column_letter] = cell_length
                except:
                    pass

        # Apply widths
        for column_letter, width in column_widths.items():
            adjusted_width = min(max(width + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _add_alternating_rows(self, worksheet):
        """Th√™m m√†u xen k·∫Ω cho c√°c h√†ng"""
        light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        for row_num in range(3, worksheet.max_row + 1, 2):  # Start from row 3, every other row
            for cell in worksheet[row_num]:
                if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                    cell.fill = light_fill

    def _add_summary_section(self, worksheet, data: List[Dict], report_type: str):
        """Th√™m ph·∫ßn t√≥m t·∫Øt v√†o worksheet"""
        # Add summary after data
        summary_start_row = worksheet.max_row + 3

        # Summary header
        summary_cell = worksheet.cell(row=summary_start_row, column=1, value="T·ªîNG K·∫æT")
        summary_cell.font = Font(name='Arial', size=14, bold=True, color='366092')

        if report_type == "inventory":
            # Inventory summary
            total_items = len(data)
            total_quantity = sum(item.get('S·ªë L∆∞·ª£ng (kg)', 0) for item in data)

            status_counts = {}
            for item in data:
                status = item.get('Tr·∫°ng Th√°i', 'Unknown')
                status_counts[status] = status_counts.get(status, 0) + 1

            # Add summary data
            worksheet.cell(row=summary_start_row + 1, column=1, value=f"T·ªïng s·ªë m·ª•c: {total_items}")
            worksheet.cell(row=summary_start_row + 2, column=1, value=f"T·ªïng kh·ªëi l∆∞·ª£ng: {total_quantity:,.2f} kg")

            row_offset = 3
            for status, count in status_counts.items():
                worksheet.cell(row=summary_start_row + row_offset, column=1, value=f"{status}: {count} m·ª•c")
                row_offset += 1

        elif report_type == "formula":
            # Formula summary
            total_ingredients = len(data)
            total_percentage = sum(item.get('T·ª∑ L·ªá (%)', 0) for item in data)

            worksheet.cell(row=summary_start_row + 1, column=1, value=f"T·ªïng s·ªë nguy√™n li·ªáu: {total_ingredients}")
            worksheet.cell(row=summary_start_row + 2, column=1, value=f"T·ªïng t·ª∑ l·ªá: {total_percentage:.1f}%")

    def export_inventory_report_optimized(self, include_feed: bool = True, include_mix: bool = True,
                                        progress_callback=None) -> Tuple[bool, str]:
        """Xu·∫•t b√°o c√°o t·ªìn kho ƒë∆∞·ª£c t·ªëi ∆∞u h√≥a"""
        try:
            start_time = time.time()

            if not include_feed and not include_mix:
                return False, "Vui l√≤ng ch·ªçn √≠t nh·∫•t m·ªôt lo·∫°i kho ƒë·ªÉ xu·∫•t"

            if progress_callback:
                progress_callback(10, "ƒêang t·∫£i d·ªØ li·ªáu...")

            # Load real data from data managers
            feed_data = {}
            mix_data = {}

            if include_feed:
                feed_data = self._get_real_inventory_data("feed")

            if include_mix:
                mix_data = self._get_real_inventory_data("mix")

            if progress_callback:
                progress_callback(30, "ƒêang x·ª≠ l√Ω d·ªØ li·ªáu...")

            # Process data efficiently
            export_data = []

            # Process feed data from real inventory manager
            if include_feed and feed_data:
                self._process_inventory_data(export_data, feed_data, 'Kho C√°m', 15000)

            # Process mix data from real inventory manager
            if include_mix and mix_data:
                self._process_inventory_data(export_data, mix_data, 'Kho Mix', 25000)

            if not export_data:
                return False, "Kh√¥ng c√≥ d·ªØ li·ªáu t·ªìn kho ƒë·ªÉ xu·∫•t"

            if progress_callback:
                progress_callback(50, "ƒêang t·∫°o file Excel...")

            # Generate filename with options
            options = {'include_feed': include_feed, 'include_mix': include_mix}
            filename = self._generate_filename("bao_cao_ton_kho_toi_uu", options)
            file_path = self.exports_dir / filename

            # Create Excel with advanced formatting
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "B√°o C√°o T·ªìn Kho"

            # Add data to worksheet
            df = pd.DataFrame(export_data)
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)

            if progress_callback:
                progress_callback(70, "ƒêang ƒë·ªãnh d·∫°ng Excel...")

            # Apply advanced formatting
            self._apply_advanced_formatting(worksheet, f"A1:F{len(export_data)+1}", "inventory")

            # Add summary section
            self._add_summary_section(worksheet, export_data, "inventory")

            if progress_callback:
                progress_callback(90, "ƒêang l∆∞u file...")

            # Save workbook
            workbook.save(file_path)

            end_time = time.time()
            processing_time = round(end_time - start_time, 2)

            return True, f"Xu·∫•t b√°o c√°o t·ªìn kho t·ªëi ∆∞u th√†nh c√¥ng!\nFile: {filename}\nV·ªã tr√≠: {file_path}\nS·ªë m·ª•c: {len(export_data)}\nTh·ªùi gian x·ª≠ l√Ω: {processing_time}s"

        except Exception as e:
            return False, f"L·ªói xu·∫•t b√°o c√°o t·ªìn kho: {str(e)}"

    def export_formula_report_optimized(self, include_feed: bool = True, include_mix: bool = True,
                                      progress_callback=None) -> Tuple[bool, str]:
        """Xu·∫•t b√°o c√°o c√¥ng th·ª©c ƒë∆∞·ª£c t·ªëi ∆∞u h√≥a"""
        try:
            start_time = time.time()

            if not include_feed and not include_mix:
                return False, "Vui l√≤ng ch·ªçn √≠t nh·∫•t m·ªôt lo·∫°i c√¥ng th·ª©c ƒë·ªÉ xu·∫•t"

            if progress_callback:
                progress_callback(10, "ƒêang t·∫£i c√¥ng th·ª©c...")

            # Load real formula data from formula manager
            feed_formula = {}
            mix_formula = {}

            if include_feed:
                feed_formula = self._get_real_formula_data("feed")

            if include_mix:
                mix_formula = self._get_real_formula_data("mix")

            if progress_callback:
                progress_callback(30, "ƒêang t·∫£i d·ªØ li·ªáu t·ªìn kho...")

            # Load real inventory data for availability check
            feed_inventory = self._get_real_inventory_data("feed") if include_feed else {}
            mix_inventory = self._get_real_inventory_data("mix") if include_mix else {}

            if progress_callback:
                progress_callback(50, "ƒêang x·ª≠ l√Ω d·ªØ li·ªáu c√¥ng th·ª©c...")

            export_data = []

            # Process feed formula
            if include_feed and feed_formula:
                for ingredient, percentage in feed_formula.items():
                    current_stock = feed_inventory.get(ingredient, 0)
                    usage_per_batch = (percentage / 100) * 1000  # For 1000kg batch
                    possible_batches = int(current_stock / usage_per_batch) if usage_per_batch > 0 else 0

                    export_data.append({
                        'Lo·∫°i C√¥ng Th·ª©c': 'C√¥ng Th·ª©c C√°m',
                        'Nguy√™n Li·ªáu': ingredient,
                        'T·ª∑ L·ªá (%)': percentage,
                        'Kh·ªëi L∆∞·ª£ng/1000kg': round(usage_per_batch, 2),
                        'T·ªìn Kho Hi·ªán T·∫°i (kg)': current_stock,
                        'S·ªë L√¥ C√≥ Th·ªÉ S·∫£n Xu·∫•t': possible_batches,
                        'Tr·∫°ng Th√°i Nguy√™n Li·ªáu': self._get_stock_status(current_stock),
                        'Ghi Ch√∫': 'Nguy√™n li·ªáu ch√≠nh' if percentage >= 20 else 'Nguy√™n li·ªáu ph·ª•'
                    })

            # Process mix formula
            if include_mix and mix_formula:
                for ingredient, percentage in mix_formula.items():
                    current_stock = mix_inventory.get(ingredient, 0)
                    usage_per_batch = (percentage / 100) * 1000
                    possible_batches = int(current_stock / usage_per_batch) if usage_per_batch > 0 else 0

                    export_data.append({
                        'Lo·∫°i C√¥ng Th·ª©c': 'C√¥ng Th·ª©c Mix',
                        'Nguy√™n Li·ªáu': ingredient,
                        'T·ª∑ L·ªá (%)': percentage,
                        'Kh·ªëi L∆∞·ª£ng/1000kg': round(usage_per_batch, 2),
                        'T·ªìn Kho Hi·ªán T·∫°i (kg)': current_stock,
                        'S·ªë L√¥ C√≥ Th·ªÉ S·∫£n Xu·∫•t': possible_batches,
                        'Tr·∫°ng Th√°i Nguy√™n Li·ªáu': self._get_stock_status(current_stock),
                        'Ghi Ch√∫': 'Nguy√™n li·ªáu ch√≠nh' if percentage >= 20 else 'Nguy√™n li·ªáu ph·ª•'
                    })

            if not export_data:
                return False, "Kh√¥ng c√≥ d·ªØ li·ªáu c√¥ng th·ª©c ƒë·ªÉ xu·∫•t"

            if progress_callback:
                progress_callback(70, "ƒêang t·∫°o file Excel...")

            # Generate filename
            options = {'include_feed': include_feed, 'include_mix': include_mix}
            filename = self._generate_filename("bao_cao_cong_thuc_toi_uu", options)
            file_path = self.exports_dir / filename

            # Create Excel with advanced formatting
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "B√°o C√°o C√¥ng Th·ª©c"

            # Add data
            df = pd.DataFrame(export_data)
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)

            if progress_callback:
                progress_callback(85, "ƒêang ƒë·ªãnh d·∫°ng Excel...")

            # Apply advanced formatting
            self._apply_advanced_formatting(worksheet, f"A1:H{len(export_data)+1}", "formula")

            # Add summary section
            self._add_summary_section(worksheet, export_data, "formula")

            # Add production capacity analysis
            self._add_production_analysis(worksheet, export_data)

            if progress_callback:
                progress_callback(95, "ƒêang l∆∞u file...")

            workbook.save(file_path)

            end_time = time.time()
            processing_time = round(end_time - start_time, 2)

            return True, f"Xu·∫•t b√°o c√°o c√¥ng th·ª©c t·ªëi ∆∞u th√†nh c√¥ng!\nFile: {filename}\nV·ªã tr√≠: {file_path}\nS·ªë m·ª•c: {len(export_data)}\nTh·ªùi gian x·ª≠ l√Ω: {processing_time}s"

        except Exception as e:
            return False, f"L·ªói xu·∫•t b√°o c√°o c√¥ng th·ª©c: {str(e)}"

    def _add_production_analysis(self, worksheet, data: List[Dict]):
        """Th√™m ph√¢n t√≠ch kh·∫£ nƒÉng s·∫£n xu·∫•t"""
        analysis_start_row = worksheet.max_row + 2

        # Analysis header
        analysis_cell = worksheet.cell(row=analysis_start_row, column=1, value="PH√ÇN T√çCH KH·∫¢ NƒÇNG S·∫¢N XU·∫§T")
        analysis_cell.font = Font(name='Arial', size=12, bold=True, color='D35400')

        # Calculate minimum production capacity
        min_batches = float('inf')
        limiting_ingredient = ""

        for item in data:
            batches = item.get('S·ªë L√¥ C√≥ Th·ªÉ S·∫£n Xu·∫•t', 0)
            if batches < min_batches:
                min_batches = batches
                limiting_ingredient = item.get('Nguy√™n Li·ªáu', '')

        if min_batches == float('inf'):
            min_batches = 0

        worksheet.cell(row=analysis_start_row + 1, column=1,
                      value=f"S·ªë l√¥ t·ªëi ƒëa c√≥ th·ªÉ s·∫£n xu·∫•t: {min_batches}")
        worksheet.cell(row=analysis_start_row + 2, column=1,
                      value=f"Nguy√™n li·ªáu h·∫°n ch·∫ø: {limiting_ingredient}")
        worksheet.cell(row=analysis_start_row + 3, column=1,
                      value=f"Kh·ªëi l∆∞·ª£ng t·ªëi ƒëa: {min_batches * 1000:,.0f} kg")

    def export_summary_report_optimized(self, progress_callback=None) -> Tuple[bool, str]:
        """Xu·∫•t b√°o c√°o t·ªïng h·ª£p ƒë∆∞·ª£c t·ªëi ∆∞u h√≥a"""
        try:
            start_time = time.time()

            if progress_callback:
                progress_callback(10, "ƒêang t·∫£i t·∫•t c·∫£ d·ªØ li·ªáu...")

            # Load all data concurrently
            with ThreadPoolExecutor(max_workers=4) as executor:
                futures = {
                    'feed_inventory': executor.submit(self._load_json_cached, self.config_dir / "feed_inventory.json"),
                    'mix_inventory': executor.submit(self._load_json_cached, self.config_dir / "mix_inventory.json"),
                    'feed_formula': executor.submit(self._load_json_cached, self.config_dir / "feed_formula.json"),
                    'mix_formula': executor.submit(self._load_json_cached, self.config_dir / "mix_formula.json")
                }

                # Collect results
                results = {key: future.result() for key, future in futures.items()}

            if progress_callback:
                progress_callback(40, "ƒêang t√≠nh to√°n th·ªëng k√™...")

            # Calculate comprehensive statistics
            summary_data = []

            # Inventory statistics
            feed_items = len(results['feed_inventory'])
            feed_quantity = sum(results['feed_inventory'].values())
            feed_value = feed_quantity * 15000

            mix_items = len(results['mix_inventory'])
            mix_quantity = sum(results['mix_inventory'].values())
            mix_value = mix_quantity * 25000

            # Status distribution
            feed_status_dist = self._calculate_status_distribution(results['feed_inventory'])
            mix_status_dist = self._calculate_status_distribution(results['mix_inventory'])

            # Add comprehensive summary data
            summary_data.extend([
                {'Lo·∫°i Th·ªëng K√™': 'T·ªìn Kho C√°m', 'S·ªë M·ª•c': feed_items, 'T·ªïng Kh·ªëi L∆∞·ª£ng (kg)': feed_quantity,
                 'Gi√° Tr·ªã ∆Ø·ªõc T√≠nh (VND)': feed_value, 'Ghi Ch√∫': 'Nguy√™n li·ªáu th√¥'},
                {'Lo·∫°i Th·ªëng K√™': 'T·ªìn Kho Mix', 'S·ªë M·ª•c': mix_items, 'T·ªïng Kh·ªëi L∆∞·ª£ng (kg)': mix_quantity,
                 'Gi√° Tr·ªã ∆Ø·ªõc T√≠nh (VND)': mix_value, 'Ghi Ch√∫': 'Ph·ª• gia dinh d∆∞·ª°ng'},
                {'Lo·∫°i Th·ªëng K√™': 'C√¥ng Th·ª©c C√°m', 'S·ªë M·ª•c': len(results['feed_formula']),
                 'T·ªïng Kh·ªëi L∆∞·ª£ng (kg)': sum(results['feed_formula'].values()), 'Gi√° Tr·ªã ∆Ø·ªõc T√≠nh (VND)': 0,
                 'Ghi Ch√∫': 'T·ª∑ l·ªá ph·∫ßn trƒÉm'},
                {'Lo·∫°i Th·ªëng K√™': 'C√¥ng Th·ª©c Mix', 'S·ªë M·ª•c': len(results['mix_formula']),
                 'T·ªïng Kh·ªëi L∆∞·ª£ng (kg)': sum(results['mix_formula'].values()), 'Gi√° Tr·ªã ∆Ø·ªõc T√≠nh (VND)': 0,
                 'Ghi Ch√∫': 'T·ª∑ l·ªá ph·∫ßn trƒÉm'}
            ])

            if progress_callback:
                progress_callback(70, "ƒêang t·∫°o file Excel...")

            filename = self._generate_filename("bao_cao_tong_hop_toi_uu")
            file_path = self.exports_dir / filename

            # Create comprehensive Excel report
            workbook = Workbook()

            # Main summary sheet
            summary_sheet = workbook.active
            summary_sheet.title = "T·ªïng Quan"

            df = pd.DataFrame(summary_data)
            for r in dataframe_to_rows(df, index=False, header=True):
                summary_sheet.append(r)

            self._apply_advanced_formatting(summary_sheet, f"A1:E{len(summary_data)+1}", "summary")

            # Add detailed status analysis
            self._add_detailed_status_analysis(workbook, feed_status_dist, mix_status_dist)

            if progress_callback:
                progress_callback(90, "ƒêang ho√†n t·∫•t...")

            workbook.save(file_path)

            end_time = time.time()
            processing_time = round(end_time - start_time, 2)

            return True, f"Xu·∫•t b√°o c√°o t·ªïng h·ª£p t·ªëi ∆∞u th√†nh c√¥ng!\nFile: {filename}\nV·ªã tr√≠: {file_path}\nTh·ªùi gian x·ª≠ l√Ω: {processing_time}s"

        except Exception as e:
            return False, f"L·ªói xu·∫•t b√°o c√°o t·ªïng h·ª£p: {str(e)}"

    def _calculate_status_distribution(self, inventory_data: Dict) -> Dict:
        """T√≠nh ph√¢n b·ªë tr·∫°ng th√°i t·ªìn kho"""
        status_dist = {'ƒê·ªß': 0, '√çt': 0, 'S·∫Øp h·∫øt': 0, 'H·∫øt h√†ng': 0}

        for quantity in inventory_data.values():
            status = self._get_stock_status(quantity)
            status_dist[status] += 1

        return status_dist

    def _add_detailed_status_analysis(self, workbook, feed_status: Dict, mix_status: Dict):
        """Th√™m sheet ph√¢n t√≠ch tr·∫°ng th√°i chi ti·∫øt"""
        status_sheet = workbook.create_sheet("Ph√¢n T√≠ch Tr·∫°ng Th√°i")

        # Headers
        status_sheet.append(['Lo·∫°i Kho', 'Tr·∫°ng Th√°i', 'S·ªë L∆∞·ª£ng M·ª•c', 'T·ª∑ L·ªá (%)'])

        # Feed status data
        total_feed = sum(feed_status.values())
        for status, count in feed_status.items():
            percentage = (count / total_feed * 100) if total_feed > 0 else 0
            status_sheet.append(['Kho C√°m', status, count, round(percentage, 1)])

        # Mix status data
        total_mix = sum(mix_status.values())
        for status, count in mix_status.items():
            percentage = (count / total_mix * 100) if total_mix > 0 else 0
            status_sheet.append(['Kho Mix', status, count, round(percentage, 1)])

        # Apply formatting
        self._apply_advanced_formatting(status_sheet, f"A1:D{status_sheet.max_row}", "status")

    def get_export_directory(self) -> str:
        """L·∫•y ƒë∆∞·ªùng d·∫´n th∆∞ m·ª•c xu·∫•t"""
        return str(self.exports_dir.absolute())

    def list_exported_files(self) -> List[str]:
        """Li·ªát k√™ c√°c file ƒë√£ xu·∫•t"""
        try:
            excel_files = list(self.exports_dir.glob("*.xlsx"))
            return [f.name for f in sorted(excel_files, key=lambda x: x.stat().st_mtime, reverse=True)]
        except:
            return []

    def clear_cache(self):
        """X√≥a cache ƒë·ªÉ refresh d·ªØ li·ªáu"""
        self._clear_cache()

    def _ensure_daily_consumption_data(self):
        """ƒê·∫£m b·∫£o c√≥ d·ªØ li·ªáu ti√™u th·ª• h√†ng ng√†y"""
        try:
            # Ki·ªÉm tra xem c√≥ file d·ªØ li·ªáu n√†o kh√¥ng
            daily_files = list(self.daily_consumption_dir.glob("daily_consumption_*.json"))

            if not daily_files:
                # T·∫°o d·ªØ li·ªáu m·∫´u n·∫øu ch∆∞a c√≥
                from src.data.daily_consumption_data import DailyConsumptionDataGenerator
                generator = DailyConsumptionDataGenerator()
                generator.save_sample_data(30)
                print("ƒê√£ t·∫°o d·ªØ li·ªáu ti√™u th·ª• h√†ng ng√†y m·∫´u")
        except Exception as e:
            print(f"L·ªói t·∫°o d·ªØ li·ªáu ti√™u th·ª• h√†ng ng√†y: {e}")

    def _load_daily_consumption_data(self, start_date: datetime, end_date: datetime) -> Dict:
        """T·∫£i d·ªØ li·ªáu ti√™u th·ª• h√†ng ng√†y trong kho·∫£ng th·ªùi gian"""
        daily_data = {}

        # T·∫°o danh s√°ch c√°c th√°ng c·∫ßn t·∫£i
        current_date = start_date.replace(day=1)  # ƒê·∫ßu th√°ng
        end_month = end_date.replace(day=1)

        while current_date <= end_month:
            month_key = current_date.strftime("%Y-%m")
            file_path = self.daily_consumption_dir / f"daily_consumption_{month_key}.json"

            if file_path.exists():
                try:
                    month_data = self._load_json_cached(file_path)

                    # L·ªçc d·ªØ li·ªáu theo kho·∫£ng th·ªùi gian
                    for date_str, day_data in month_data.items():
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                        if start_date <= date_obj <= end_date:
                            daily_data[date_str] = day_data

                except Exception as e:
                    print(f"L·ªói t·∫£i d·ªØ li·ªáu th√°ng {month_key}: {e}")

            # Chuy·ªÉn sang th√°ng ti·∫øp theo
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)

        return daily_data

    def export_daily_regional_report(self, start_date: datetime, end_date: datetime,
                                   selected_regions: List[str] = None,
                                   include_feed: bool = True, include_mix: bool = True,
                                   progress_callback=None) -> Tuple[bool, str]:
        """Xu·∫•t b√°o c√°o ti√™u th·ª• h√†ng ng√†y theo khu v·ª±c"""
        try:
            start_time = time.time()

            if progress_callback:
                progress_callback(10, "ƒêang t·∫£i d·ªØ li·ªáu ti√™u th·ª• h√†ng ng√†y...")

            # T·∫£i d·ªØ li·ªáu ti√™u th·ª•
            daily_data = self._load_daily_consumption_data(start_date, end_date)

            if not daily_data:
                return False, "Kh√¥ng c√≥ d·ªØ li·ªáu ti√™u th·ª• trong kho·∫£ng th·ªùi gian ƒë√£ ch·ªçn"

            if progress_callback:
                progress_callback(30, "ƒêang x·ª≠ l√Ω d·ªØ li·ªáu theo khu v·ª±c...")

            # X·ª≠ l√Ω d·ªØ li·ªáu theo khu v·ª±c
            processed_data = self._process_regional_data(daily_data, selected_regions, include_feed, include_mix)

            if progress_callback:
                progress_callback(50, "ƒêang t·∫°o file Excel...")

            # T·∫°o filename
            date_range = f"{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}"
            regions_suffix = "_".join(selected_regions) if selected_regions else "all_regions"
            filename = f"bao_cao_hang_ngay_{date_range}_{regions_suffix}.xlsx"
            file_path = self.exports_dir / filename

            # T·∫°o Excel v·ªõi multiple sheets
            workbook = Workbook()

            # Sheet t·ªïng quan
            self._create_daily_overview_sheet(workbook, processed_data, start_date, end_date)

            if progress_callback:
                progress_callback(70, "ƒêang t·∫°o sheet chi ti·∫øt...")

            # Sheet chi ti·∫øt theo khu v·ª±c
            if selected_regions:
                for region_id in selected_regions:
                    if region_id in processed_data['regions']:
                        self._create_regional_detail_sheet(workbook, region_id, processed_data['regions'][region_id])

            # Sheet ph√¢n t√≠ch xu h∆∞·ªõng
            self._create_trend_analysis_sheet(workbook, processed_data)

            if progress_callback:
                progress_callback(90, "ƒêang l∆∞u file...")

            workbook.save(file_path)

            end_time = time.time()
            processing_time = round(end_time - start_time, 2)

            return True, f"Xu·∫•t b√°o c√°o h√†ng ng√†y th√†nh c√¥ng!\nFile: {filename}\nV·ªã tr√≠: {file_path}\nS·ªë ng√†y: {len(daily_data)}\nTh·ªùi gian x·ª≠ l√Ω: {processing_time}s"

        except Exception as e:
            return False, f"L·ªói xu·∫•t b√°o c√°o h√†ng ng√†y: {str(e)}"

    def export_feed_component_report(self, start_date: datetime, end_date: datetime,
                                   selected_regions: List[str] = None,
                                   progress_callback=None) -> Tuple[bool, str]:
        """Xu·∫•t b√°o c√°o chi ti·∫øt th√†nh ph·∫ßn c√°m"""
        try:
            start_time = time.time()

            if progress_callback:
                progress_callback(10, "ƒêang t·∫£i d·ªØ li·ªáu th√†nh ph·∫ßn c√°m...")

            daily_data = self._load_daily_consumption_data(start_date, end_date)

            if not daily_data:
                return False, "Kh√¥ng c√≥ d·ªØ li·ªáu th√†nh ph·∫ßn c√°m trong kho·∫£ng th·ªùi gian ƒë√£ ch·ªçn"

            if progress_callback:
                progress_callback(40, "ƒêang ph√¢n t√≠ch th√†nh ph·∫ßn c√°m...")

            # X·ª≠ l√Ω d·ªØ li·ªáu th√†nh ph·∫ßn c√°m
            feed_analysis = self._analyze_feed_components(daily_data, selected_regions)

            if progress_callback:
                progress_callback(60, "ƒêang t·∫°o b√°o c√°o Excel...")

            # T·∫°o filename
            date_range = f"{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}"
            regions_suffix = "_".join(selected_regions) if selected_regions else "all_regions"
            filename = f"bao_cao_thanh_phan_cam_{date_range}_{regions_suffix}.xlsx"
            file_path = self.exports_dir / filename

            # T·∫°o Excel v·ªõi ph√¢n t√≠ch chi ti·∫øt
            workbook = Workbook()

            # Sheet t·ªïng quan th√†nh ph·∫ßn
            self._create_feed_component_overview_sheet(workbook, feed_analysis)

            # Sheet xu h∆∞·ªõng ti√™u th·ª•
            self._create_feed_consumption_trend_sheet(workbook, feed_analysis)

            # Sheet so s√°nh v·ªõi c√¥ng th·ª©c chu·∫©n
            self._create_feed_formula_comparison_sheet(workbook, feed_analysis)

            if progress_callback:
                progress_callback(90, "ƒêang l∆∞u file...")

            workbook.save(file_path)

            end_time = time.time()
            processing_time = round(end_time - start_time, 2)

            return True, f"Xu·∫•t b√°o c√°o th√†nh ph·∫ßn c√°m th√†nh c√¥ng!\nFile: {filename}\nV·ªã tr√≠: {file_path}\nTh·ªùi gian x·ª≠ l√Ω: {processing_time}s"

        except Exception as e:
            return False, f"L·ªói xu·∫•t b√°o c√°o th√†nh ph·∫ßn c√°m: {str(e)}"

    def export_mix_component_report(self, start_date: datetime, end_date: datetime,
                                  selected_regions: List[str] = None,
                                  progress_callback=None) -> Tuple[bool, str]:
        """Xu·∫•t b√°o c√°o chi ti·∫øt th√†nh ph·∫ßn mix"""
        try:
            start_time = time.time()

            if progress_callback:
                progress_callback(10, "ƒêang t·∫£i d·ªØ li·ªáu th√†nh ph·∫ßn mix...")

            daily_data = self._load_daily_consumption_data(start_date, end_date)

            if not daily_data:
                return False, "Kh√¥ng c√≥ d·ªØ li·ªáu th√†nh ph·∫ßn mix trong kho·∫£ng th·ªùi gian ƒë√£ ch·ªçn"

            if progress_callback:
                progress_callback(40, "ƒêang ph√¢n t√≠ch th√†nh ph·∫ßn mix...")

            # X·ª≠ l√Ω d·ªØ li·ªáu th√†nh ph·∫ßn mix
            mix_analysis = self._analyze_mix_components(daily_data, selected_regions)

            if progress_callback:
                progress_callback(60, "ƒêang t·∫°o b√°o c√°o Excel...")

            # T·∫°o filename
            date_range = f"{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}"
            regions_suffix = "_".join(selected_regions) if selected_regions else "all_regions"
            filename = f"bao_cao_thanh_phan_mix_{date_range}_{regions_suffix}.xlsx"
            file_path = self.exports_dir / filename

            # T·∫°o Excel v·ªõi ph√¢n t√≠ch chi ti·∫øt
            workbook = Workbook()

            # Sheet t·ªïng quan th√†nh ph·∫ßn
            self._create_mix_component_overview_sheet(workbook, mix_analysis)

            # Sheet xu h∆∞·ªõng ti√™u th·ª•
            self._create_mix_consumption_trend_sheet(workbook, mix_analysis)

            # Sheet ph√¢n t√≠ch hi·ªáu qu·∫£
            self._create_mix_efficiency_analysis_sheet(workbook, mix_analysis)

            if progress_callback:
                progress_callback(90, "ƒêang l∆∞u file...")

            workbook.save(file_path)

            end_time = time.time()
            processing_time = round(end_time - start_time, 2)

            return True, f"Xu·∫•t b√°o c√°o th√†nh ph·∫ßn mix th√†nh c√¥ng!\nFile: {filename}\nV·ªã tr√≠: {file_path}\nTh·ªùi gian x·ª≠ l√Ω: {processing_time}s"

        except Exception as e:
            return False, f"L·ªói xu·∫•t b√°o c√°o th√†nh ph·∫ßn mix: {str(e)}"

    def _process_regional_data(self, daily_data: Dict, selected_regions: List[str],
                             include_feed: bool, include_mix: bool) -> Dict:
        """X·ª≠ l√Ω d·ªØ li·ªáu theo khu v·ª±c"""
        processed_data = {
            'summary': {
                'total_days': len(daily_data),
                'date_range': {
                    'start': min(daily_data.keys()),
                    'end': max(daily_data.keys())
                },
                'total_production': 0,
                'total_feed_consumption': {},
                'total_mix_consumption': {}
            },
            'regions': {},
            'daily_totals': {}
        }

        # Import region definitions
        try:
            from src.data.daily_consumption_data import REGIONS
        except ImportError:
            REGIONS = {
                "mien_bac": {"name": "Mi·ªÅn B·∫Øc"},
                "mien_trung": {"name": "Mi·ªÅn Trung"},
                "mien_nam": {"name": "Mi·ªÅn Nam"}
            }

        # X·ª≠ l√Ω t·ª´ng ng√†y
        for date_str, day_data in daily_data.items():
            daily_total = {
                'date': date_str,
                'total_production': 0,
                'regions_production': {}
            }

            # X·ª≠ l√Ω t·ª´ng khu v·ª±c
            for region_id, region_data in day_data.get('regions', {}).items():
                if selected_regions and region_id not in selected_regions:
                    continue

                # Kh·ªüi t·∫°o d·ªØ li·ªáu khu v·ª±c n·∫øu ch∆∞a c√≥
                if region_id not in processed_data['regions']:
                    processed_data['regions'][region_id] = {
                        'region_name': region_data.get('region_name', REGIONS.get(region_id, {}).get('name', region_id)),
                        'total_production': 0,
                        'daily_production': {},
                        'feed_consumption': {},
                        'mix_consumption': {},
                        'animal_distribution': {}
                    }

                region_production = region_data.get('total_production', 0)
                processed_data['regions'][region_id]['total_production'] += region_production
                processed_data['regions'][region_id]['daily_production'][date_str] = region_production

                daily_total['total_production'] += region_production
                daily_total['regions_production'][region_id] = region_production

                # X·ª≠ l√Ω ti√™u th·ª• c√°m
                if include_feed:
                    feed_consumption = region_data.get('feed_consumption', {})
                    for component, amount in feed_consumption.items():
                        if component not in processed_data['regions'][region_id]['feed_consumption']:
                            processed_data['regions'][region_id]['feed_consumption'][component] = 0
                        processed_data['regions'][region_id]['feed_consumption'][component] += amount

                        if component not in processed_data['summary']['total_feed_consumption']:
                            processed_data['summary']['total_feed_consumption'][component] = 0
                        processed_data['summary']['total_feed_consumption'][component] += amount

                # X·ª≠ l√Ω ti√™u th·ª• mix
                if include_mix:
                    mix_consumption = region_data.get('mix_consumption', {})
                    for component, amount in mix_consumption.items():
                        if component not in processed_data['regions'][region_id]['mix_consumption']:
                            processed_data['regions'][region_id]['mix_consumption'][component] = 0
                        processed_data['regions'][region_id]['mix_consumption'][component] += amount

                        if component not in processed_data['summary']['total_mix_consumption']:
                            processed_data['summary']['total_mix_consumption'][component] = 0
                        processed_data['summary']['total_mix_consumption'][component] += amount

            processed_data['daily_totals'][date_str] = daily_total
            processed_data['summary']['total_production'] += daily_total['total_production']

        return processed_data

    def _analyze_feed_components(self, daily_data: Dict, selected_regions: List[str]) -> Dict:
        """Ph√¢n t√≠ch chi ti·∫øt th√†nh ph·∫ßn c√°m"""
        try:
            from src.data.daily_consumption_data import FEED_COMPONENTS, ANIMAL_TYPES
        except ImportError:
            FEED_COMPONENTS = {}
            ANIMAL_TYPES = {}

        analysis = {
            'components_summary': {},
            'daily_consumption': {},
            'regional_breakdown': {},
            'formula_comparison': {},
            'cost_analysis': {}
        }

        # Ph√¢n t√≠ch t·ª´ng th√†nh ph·∫ßn
        for date_str, day_data in daily_data.items():
            analysis['daily_consumption'][date_str] = {}

            for region_id, region_data in day_data.get('regions', {}).items():
                if selected_regions and region_id not in selected_regions:
                    continue

                feed_consumption = region_data.get('feed_consumption', {})

                for component, amount in feed_consumption.items():
                    # T·ªïng h·ª£p theo th√†nh ph·∫ßn
                    if component not in analysis['components_summary']:
                        analysis['components_summary'][component] = {
                            'total_consumption': 0,
                            'average_daily': 0,
                            'max_daily': 0,
                            'min_daily': float('inf'),
                            'component_info': FEED_COMPONENTS.get(component, {})
                        }

                    analysis['components_summary'][component]['total_consumption'] += amount

                    # Theo ng√†y
                    if component not in analysis['daily_consumption'][date_str]:
                        analysis['daily_consumption'][date_str][component] = 0
                    analysis['daily_consumption'][date_str][component] += amount

                    # Theo khu v·ª±c
                    if region_id not in analysis['regional_breakdown']:
                        analysis['regional_breakdown'][region_id] = {}
                    if component not in analysis['regional_breakdown'][region_id]:
                        analysis['regional_breakdown'][region_id][component] = 0
                    analysis['regional_breakdown'][region_id][component] += amount

        # T√≠nh to√°n th·ªëng k√™
        total_days = len(daily_data)
        for component, data in analysis['components_summary'].items():
            data['average_daily'] = data['total_consumption'] / total_days if total_days > 0 else 0

            # T√¨m min/max daily
            daily_amounts = []
            for day_consumption in analysis['daily_consumption'].values():
                daily_amounts.append(day_consumption.get(component, 0))

            if daily_amounts:
                data['max_daily'] = max(daily_amounts)
                data['min_daily'] = min(daily_amounts)

            # T√≠nh chi ph√≠
            price_per_kg = data['component_info'].get('price_per_kg', 0)
            data['total_cost'] = data['total_consumption'] * price_per_kg
            data['average_daily_cost'] = data['average_daily'] * price_per_kg

        return analysis

    def _analyze_mix_components(self, daily_data: Dict, selected_regions: List[str]) -> Dict:
        """Ph√¢n t√≠ch chi ti·∫øt th√†nh ph·∫ßn mix"""
        try:
            from src.data.daily_consumption_data import MIX_COMPONENTS
        except ImportError:
            MIX_COMPONENTS = {}

        analysis = {
            'components_summary': {},
            'daily_consumption': {},
            'regional_breakdown': {},
            'efficiency_analysis': {},
            'cost_analysis': {}
        }

        # Ph√¢n t√≠ch t∆∞∆°ng t·ª± nh∆∞ feed components
        for date_str, day_data in daily_data.items():
            analysis['daily_consumption'][date_str] = {}

            for region_id, region_data in day_data.get('regions', {}).items():
                if selected_regions and region_id not in selected_regions:
                    continue

                mix_consumption = region_data.get('mix_consumption', {})

                for component, amount in mix_consumption.items():
                    # T·ªïng h·ª£p theo th√†nh ph·∫ßn
                    if component not in analysis['components_summary']:
                        analysis['components_summary'][component] = {
                            'total_consumption': 0,
                            'average_daily': 0,
                            'max_daily': 0,
                            'min_daily': float('inf'),
                            'component_info': MIX_COMPONENTS.get(component, {})
                        }

                    analysis['components_summary'][component]['total_consumption'] += amount

                    # Theo ng√†y
                    if component not in analysis['daily_consumption'][date_str]:
                        analysis['daily_consumption'][date_str][component] = 0
                    analysis['daily_consumption'][date_str][component] += amount

                    # Theo khu v·ª±c
                    if region_id not in analysis['regional_breakdown']:
                        analysis['regional_breakdown'][region_id] = {}
                    if component not in analysis['regional_breakdown'][region_id]:
                        analysis['regional_breakdown'][region_id][component] = 0
                    analysis['regional_breakdown'][region_id][component] += amount

        # T√≠nh to√°n th·ªëng k√™ v√† hi·ªáu qu·∫£
        total_days = len(daily_data)
        for component, data in analysis['components_summary'].items():
            data['average_daily'] = data['total_consumption'] / total_days if total_days > 0 else 0

            # T√¨m min/max daily
            daily_amounts = []
            for day_consumption in analysis['daily_consumption'].values():
                daily_amounts.append(day_consumption.get(component, 0))

            if daily_amounts:
                data['max_daily'] = max(daily_amounts)
                data['min_daily'] = min(daily_amounts)

            # T√≠nh chi ph√≠
            price_per_kg = data['component_info'].get('price_per_kg', 0)
            data['total_cost'] = data['total_consumption'] * price_per_kg
            data['average_daily_cost'] = data['average_daily'] * price_per_kg

            # Ph√¢n t√≠ch hi·ªáu qu·∫£ (so v·ªõi dosage range)
            dosage_range = data['component_info'].get('dosage_range', '')
            if dosage_range and '-' in dosage_range:
                try:
                    min_dosage, max_dosage = dosage_range.replace('%', '').split('-')
                    min_dosage = float(min_dosage)
                    max_dosage = float(max_dosage)

                    # T√≠nh t·ª∑ l·ªá s·ª≠ d·ª•ng trung b√¨nh (gi·∫£ ƒë·ªãnh)
                    avg_usage_rate = (data['average_daily'] / 1000) * 100  # Gi·∫£ ƒë·ªãnh 1000kg/ng√†y

                    if avg_usage_rate < min_dosage:
                        data['efficiency_status'] = 'D∆∞·ªõi m·ª©c khuy·∫øn ngh·ªã'
                    elif avg_usage_rate > max_dosage:
                        data['efficiency_status'] = 'V∆∞·ª£t m·ª©c khuy·∫øn ngh·ªã'
                    else:
                        data['efficiency_status'] = 'Trong kho·∫£ng khuy·∫øn ngh·ªã'

                    data['usage_rate'] = round(avg_usage_rate, 3)
                except:
                    data['efficiency_status'] = 'Kh√¥ng x√°c ƒë·ªãnh'

        return analysis

    def _create_daily_overview_sheet(self, workbook: Workbook, processed_data: Dict,
                                   start_date: datetime, end_date: datetime):
        """T·∫°o sheet t·ªïng quan b√°o c√°o h√†ng ng√†y"""
        # Remove default sheet and create overview
        if 'Sheet' in [ws.title for ws in workbook.worksheets]:
            workbook.remove(workbook['Sheet'])

        overview_sheet = workbook.create_sheet("T·ªïng Quan H√†ng Ng√†y")

        # Header th√¥ng tin
        overview_sheet.append(['B√ÅO C√ÅO TI√äU TH·ª§ H√ÄNG NG√ÄY'])
        overview_sheet.append([f'T·ª´ ng√†y: {start_date.strftime("%d/%m/%Y")} - ƒê·∫øn ng√†y: {end_date.strftime("%d/%m/%Y")}'])
        overview_sheet.append([f'T·ªïng s·ªë ng√†y: {processed_data["summary"]["total_days"]}'])
        overview_sheet.append([])  # Empty row

        # Th·ªëng k√™ t·ªïng quan
        overview_sheet.append(['TH·ªêNG K√ä T·ªîNG QUAN'])
        overview_sheet.append(['Ch·ªâ s·ªë', 'Gi√° tr·ªã', 'ƒê∆°n v·ªã'])
        overview_sheet.append(['T·ªïng s·∫£n l∆∞·ª£ng', processed_data['summary']['total_production'], 'kg'])
        overview_sheet.append(['S·∫£n l∆∞·ª£ng trung b√¨nh/ng√†y',
                              round(processed_data['summary']['total_production'] / processed_data['summary']['total_days'], 2), 'kg'])
        overview_sheet.append([])  # Empty row

        # Th·ªëng k√™ theo khu v·ª±c
        overview_sheet.append(['TH·ªêNG K√ä THEO KHU V·ª∞C'])
        overview_sheet.append(['Khu v·ª±c', 'T·ªïng s·∫£n l∆∞·ª£ng (kg)', 'T·ª∑ l·ªá (%)', 'Trung b√¨nh/ng√†y (kg)'])

        total_production = processed_data['summary']['total_production']
        for region_id, region_data in processed_data['regions'].items():
            region_production = region_data['total_production']
            percentage = (region_production / total_production * 100) if total_production > 0 else 0
            avg_daily = region_production / processed_data['summary']['total_days']

            overview_sheet.append([
                region_data['region_name'],
                round(region_production, 2),
                round(percentage, 1),
                round(avg_daily, 2)
            ])

        overview_sheet.append([])  # Empty row

        # Top 10 th√†nh ph·∫ßn c√°m ti√™u th·ª• nhi·ªÅu nh·∫•t
        if processed_data['summary']['total_feed_consumption']:
            overview_sheet.append(['TOP 10 TH√ÄNH PH·∫¶N C√ÅM TI√äU TH·ª§ NHI·ªÄU NH·∫§T'])
            overview_sheet.append(['Th√†nh ph·∫ßn', 'T·ªïng ti√™u th·ª• (kg)', 'Trung b√¨nh/ng√†y (kg)'])

            sorted_feed = sorted(processed_data['summary']['total_feed_consumption'].items(),
                               key=lambda x: x[1], reverse=True)[:10]

            for component, total_amount in sorted_feed:
                avg_daily = total_amount / processed_data['summary']['total_days']
                overview_sheet.append([component, round(total_amount, 2), round(avg_daily, 2)])

        # Apply formatting
        self._apply_advanced_formatting(overview_sheet, f"A1:C{overview_sheet.max_row}", "daily_overview")

    def _create_regional_detail_sheet(self, workbook: Workbook, region_id: str, region_data: Dict):
        """T·∫°o sheet chi ti·∫øt cho m·ªôt khu v·ª±c"""
        sheet_name = f"Chi Ti·∫øt {region_data['region_name']}"
        detail_sheet = workbook.create_sheet(sheet_name)

        # Header
        detail_sheet.append([f'CHI TI·∫æT KHU V·ª∞C: {region_data["region_name"].upper()}'])
        detail_sheet.append([])

        # Th√¥ng tin t·ªïng quan khu v·ª±c
        detail_sheet.append(['TH√îNG TIN T·ªîNG QUAN'])
        detail_sheet.append(['T·ªïng s·∫£n l∆∞·ª£ng:', f'{region_data["total_production"]:,.2f} kg'])
        detail_sheet.append(['S·ªë ng√†y c√≥ d·ªØ li·ªáu:', len(region_data['daily_production'])])
        detail_sheet.append([])

        # S·∫£n l∆∞·ª£ng theo ng√†y
        detail_sheet.append(['S·∫¢N L∆Ø·ª¢NG THEO NG√ÄY'])
        detail_sheet.append(['Ng√†y', 'S·∫£n l∆∞·ª£ng (kg)', 'Ghi ch√∫'])

        for date_str, production in sorted(region_data['daily_production'].items()):
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            weekday = date_obj.strftime("%A")
            detail_sheet.append([
                date_obj.strftime("%d/%m/%Y"),
                round(production, 2),
                f'Th·ª© {weekday}'
            ])

        detail_sheet.append([])

        # Ti√™u th·ª• th√†nh ph·∫ßn c√°m
        if region_data['feed_consumption']:
            detail_sheet.append(['TI√äU TH·ª§ TH√ÄNH PH·∫¶N C√ÅM'])
            detail_sheet.append(['Th√†nh ph·∫ßn', 'T·ªïng ti√™u th·ª• (kg)', 'Trung b√¨nh/ng√†y (kg)', 'T·ª∑ l·ªá (%)'])

            total_feed = sum(region_data['feed_consumption'].values())
            days_count = len(region_data['daily_production'])

            for component, amount in sorted(region_data['feed_consumption'].items(), key=lambda x: x[1], reverse=True):
                avg_daily = amount / days_count if days_count > 0 else 0
                percentage = (amount / total_feed * 100) if total_feed > 0 else 0

                detail_sheet.append([
                    component,
                    round(amount, 2),
                    round(avg_daily, 2),
                    round(percentage, 1)
                ])

        detail_sheet.append([])

        # Ti√™u th·ª• th√†nh ph·∫ßn mix
        if region_data['mix_consumption']:
            detail_sheet.append(['TI√äU TH·ª§ TH√ÄNH PH·∫¶N MIX'])
            detail_sheet.append(['Th√†nh ph·∫ßn', 'T·ªïng ti√™u th·ª• (kg)', 'Trung b√¨nh/ng√†y (kg)', 'T·ª∑ l·ªá (%)'])

            total_mix = sum(region_data['mix_consumption'].values())
            days_count = len(region_data['daily_production'])

            for component, amount in sorted(region_data['mix_consumption'].items(), key=lambda x: x[1], reverse=True):
                avg_daily = amount / days_count if days_count > 0 else 0
                percentage = (amount / total_mix * 100) if total_mix > 0 else 0

                detail_sheet.append([
                    component,
                    round(amount, 2),
                    round(avg_daily, 2),
                    round(percentage, 1)
                ])

        # Apply formatting
        self._apply_advanced_formatting(detail_sheet, f"A1:D{detail_sheet.max_row}", "regional_detail")

    def _create_trend_analysis_sheet(self, workbook: Workbook, processed_data: Dict):
        """T·∫°o sheet ph√¢n t√≠ch xu h∆∞·ªõng"""
        trend_sheet = workbook.create_sheet("Ph√¢n T√≠ch Xu H∆∞·ªõng")

        # Header
        trend_sheet.append(['PH√ÇN T√çCH XU H∆Ø·ªöNG TI√äU TH·ª§'])
        trend_sheet.append([])

        # Xu h∆∞·ªõng s·∫£n l∆∞·ª£ng theo ng√†y
        trend_sheet.append(['XU H∆Ø·ªöNG S·∫¢N L∆Ø·ª¢NG THEO NG√ÄY'])
        trend_sheet.append(['Ng√†y', 'T·ªïng s·∫£n l∆∞·ª£ng (kg)'] + [region_data['region_name'] for region_data in processed_data['regions'].values()])

        for date_str, daily_total in sorted(processed_data['daily_totals'].items()):
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            row = [date_obj.strftime("%d/%m/%Y"), round(daily_total['total_production'], 2)]

            # Th√™m s·∫£n l∆∞·ª£ng t·ª´ng khu v·ª±c
            for region_id in processed_data['regions'].keys():
                region_production = daily_total['regions_production'].get(region_id, 0)
                row.append(round(region_production, 2))

            trend_sheet.append(row)

        trend_sheet.append([])

        # Ph√¢n t√≠ch theo ng√†y trong tu·∫ßn
        trend_sheet.append(['PH√ÇN T√çCH THEO NG√ÄY TRONG TU·∫¶N'])
        trend_sheet.append(['Th·ª©', 'S·∫£n l∆∞·ª£ng trung b√¨nh (kg)', 'S·ªë ng√†y', 'T·ª∑ l·ªá (%)'])

        weekday_stats = {}
        total_production = processed_data['summary']['total_production']

        for date_str, daily_total in processed_data['daily_totals'].items():
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            weekday = date_obj.strftime("%A")

            if weekday not in weekday_stats:
                weekday_stats[weekday] = {'total': 0, 'count': 0}

            weekday_stats[weekday]['total'] += daily_total['total_production']
            weekday_stats[weekday]['count'] += 1

        weekday_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        weekday_names = ['Th·ª© 2', 'Th·ª© 3', 'Th·ª© 4', 'Th·ª© 5', 'Th·ª© 6', 'Th·ª© 7', 'Ch·ªß nh·∫≠t']

        for i, weekday in enumerate(weekday_order):
            if weekday in weekday_stats:
                stats = weekday_stats[weekday]
                avg_production = stats['total'] / stats['count']
                percentage = (stats['total'] / total_production * 100) if total_production > 0 else 0

                trend_sheet.append([
                    weekday_names[i],
                    round(avg_production, 2),
                    stats['count'],
                    round(percentage, 1)
                ])

        # Apply formatting
        self._apply_advanced_formatting(trend_sheet, f"A1:E{trend_sheet.max_row}", "trend_analysis")

    def _create_feed_component_overview_sheet(self, workbook: Workbook, feed_analysis: Dict):
        """T·∫°o sheet t·ªïng quan th√†nh ph·∫ßn c√°m"""
        if 'Sheet' in [ws.title for ws in workbook.worksheets]:
            workbook.remove(workbook['Sheet'])

        overview_sheet = workbook.create_sheet("T·ªïng Quan Th√†nh Ph·∫ßn C√°m")

        # Header
        overview_sheet.append(['B√ÅO C√ÅO CHI TI·∫æT TH√ÄNH PH·∫¶N C√ÅM'])
        overview_sheet.append([])

        # Th·ªëng k√™ t·ªïng quan
        overview_sheet.append(['TH·ªêNG K√ä T·ªîNG QUAN TH√ÄNH PH·∫¶N'])
        overview_sheet.append(['Th√†nh ph·∫ßn', 'T·ªïng ti√™u th·ª• (kg)', 'TB/ng√†y (kg)', 'Max/ng√†y (kg)',
                              'Min/ng√†y (kg)', 'Gi√°/kg (VND)', 'T·ªïng chi ph√≠ (VND)'])

        for component, data in sorted(feed_analysis['components_summary'].items(),
                                    key=lambda x: x[1]['total_consumption'], reverse=True):
            overview_sheet.append([
                component,
                round(data['total_consumption'], 2),
                round(data['average_daily'], 2),
                round(data['max_daily'], 2),
                round(data['min_daily'], 2) if data['min_daily'] != float('inf') else 0,
                data['component_info'].get('price_per_kg', 0),
                round(data.get('total_cost', 0), 0)
            ])

        overview_sheet.append([])

        # Th√¥ng tin dinh d∆∞·ª°ng
        overview_sheet.append(['TH√îNG TIN DINH D∆Ø·ª†NG TH√ÄNH PH·∫¶N'])
        overview_sheet.append(['Th√†nh ph·∫ßn', 'Protein (%)', 'NƒÉng l∆∞·ª£ng (kcal/kg)', 'Ch·∫•t x∆° (%)', 'Nh√† cung c·∫•p'])

        for component, data in feed_analysis['components_summary'].items():
            component_info = data['component_info']
            overview_sheet.append([
                component,
                component_info.get('protein', 'N/A'),
                component_info.get('energy', 'N/A'),
                component_info.get('fiber', 'N/A'),
                component_info.get('supplier', 'N/A')
            ])

        # Apply formatting
        self._apply_advanced_formatting(overview_sheet, f"A1:G{overview_sheet.max_row}", "feed_overview")

    def _create_feed_consumption_trend_sheet(self, workbook: Workbook, feed_analysis: Dict):
        """T·∫°o sheet xu h∆∞·ªõng ti√™u th·ª• c√°m"""
        trend_sheet = workbook.create_sheet("Xu H∆∞·ªõng Ti√™u Th·ª• C√°m")

        # Header
        trend_sheet.append(['XU H∆Ø·ªöNG TI√äU TH·ª§ TH√ÄNH PH·∫¶N C√ÅM THEO NG√ÄY'])
        trend_sheet.append([])

        # T·∫°o b·∫£ng d·ªØ li·ªáu theo ng√†y
        components = list(feed_analysis['components_summary'].keys())
        header = ['Ng√†y'] + components
        trend_sheet.append(header)

        for date_str in sorted(feed_analysis['daily_consumption'].keys()):
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            row = [date_obj.strftime("%d/%m/%Y")]

            daily_data = feed_analysis['daily_consumption'][date_str]
            for component in components:
                amount = daily_data.get(component, 0)
                row.append(round(amount, 2))

            trend_sheet.append(row)

        trend_sheet.append([])

        # Ph√¢n t√≠ch bi·∫øn ƒë·ªông
        trend_sheet.append(['PH√ÇN T√çCH BI·∫æN ƒê·ªòNG'])
        trend_sheet.append(['Th√†nh ph·∫ßn', 'ƒê·ªô l·ªách chu·∫©n', 'H·ªá s·ªë bi·∫øn ƒë·ªông (%)', 'Xu h∆∞·ªõng'])

        for component in components:
            daily_amounts = []
            for daily_data in feed_analysis['daily_consumption'].values():
                daily_amounts.append(daily_data.get(component, 0))

            if daily_amounts:
                import statistics
                mean_amount = statistics.mean(daily_amounts)
                std_dev = statistics.stdev(daily_amounts) if len(daily_amounts) > 1 else 0
                cv = (std_dev / mean_amount * 100) if mean_amount > 0 else 0

                # ƒê√°nh gi√° xu h∆∞·ªõng ƒë∆°n gi·∫£n
                if len(daily_amounts) >= 3:
                    first_half = daily_amounts[:len(daily_amounts)//2]
                    second_half = daily_amounts[len(daily_amounts)//2:]

                    avg_first = statistics.mean(first_half)
                    avg_second = statistics.mean(second_half)

                    if avg_second > avg_first * 1.1:
                        trend = "TƒÉng"
                    elif avg_second < avg_first * 0.9:
                        trend = "Gi·∫£m"
                    else:
                        trend = "·ªîn ƒë·ªãnh"
                else:
                    trend = "Kh√¥ng ƒë·ªß d·ªØ li·ªáu"

                trend_sheet.append([
                    component,
                    round(std_dev, 2),
                    round(cv, 1),
                    trend
                ])

        # Apply formatting
        self._apply_advanced_formatting(trend_sheet, f"A1:D{trend_sheet.max_row}", "feed_trend")

    def _create_feed_formula_comparison_sheet(self, workbook: Workbook, feed_analysis: Dict):
        """T·∫°o sheet so s√°nh v·ªõi c√¥ng th·ª©c chu·∫©n"""
        comparison_sheet = workbook.create_sheet("So S√°nh C√¥ng Th·ª©c Chu·∫©n")

        # Header
        comparison_sheet.append(['SO S√ÅNH V·ªöI C√îNG TH·ª®C CHU·∫®N'])
        comparison_sheet.append([])

        try:
            from src.data.daily_consumption_data import ANIMAL_TYPES

            # So s√°nh v·ªõi t·ª´ng lo·∫°i ƒë·ªông v·∫≠t
            for animal_type, animal_info in ANIMAL_TYPES.items():
                comparison_sheet.append([f'C√îNG TH·ª®C {animal_info["name"].upper()}'])
                comparison_sheet.append(['Th√†nh ph·∫ßn', 'C√¥ng th·ª©c chu·∫©n (%)', 'Ti√™u th·ª• th·ª±c t·∫ø (kg)',
                                       'T·ª∑ l·ªá th·ª±c t·∫ø (%)', 'Ch√™nh l·ªách (%)'])

                feed_formula = animal_info['feed_formula']
                total_actual = sum(feed_analysis['components_summary'].get(comp, {}).get('total_consumption', 0)
                                 for comp in feed_formula.keys())

                for component, standard_percentage in feed_formula.items():
                    actual_consumption = feed_analysis['components_summary'].get(component, {}).get('total_consumption', 0)
                    actual_percentage = (actual_consumption / total_actual * 100) if total_actual > 0 else 0
                    difference = actual_percentage - standard_percentage

                    comparison_sheet.append([
                        component,
                        standard_percentage,
                        round(actual_consumption, 2),
                        round(actual_percentage, 1),
                        round(difference, 1)
                    ])

                comparison_sheet.append([])  # Empty row between animal types

        except ImportError:
            comparison_sheet.append(['Kh√¥ng th·ªÉ t·∫£i d·ªØ li·ªáu c√¥ng th·ª©c chu·∫©n'])

        # Apply formatting
        self._apply_advanced_formatting(comparison_sheet, f"A1:E{comparison_sheet.max_row}", "formula_comparison")

    def _create_mix_component_overview_sheet(self, workbook: Workbook, mix_analysis: Dict):
        """T·∫°o sheet t·ªïng quan th√†nh ph·∫ßn mix"""
        if 'Sheet' in [ws.title for ws in workbook.worksheets]:
            workbook.remove(workbook['Sheet'])

        overview_sheet = workbook.create_sheet("T·ªïng Quan Th√†nh Ph·∫ßn Mix")

        # Header
        overview_sheet.append(['B√ÅO C√ÅO CHI TI·∫æT TH√ÄNH PH·∫¶N MIX'])
        overview_sheet.append([])

        # Th·ªëng k√™ t·ªïng quan
        overview_sheet.append(['TH·ªêNG K√ä T·ªîNG QUAN TH√ÄNH PH·∫¶N'])
        overview_sheet.append(['Th√†nh ph·∫ßn', 'T·ªïng ti√™u th·ª• (kg)', 'TB/ng√†y (kg)', 'Max/ng√†y (kg)',
                              'Min/ng√†y (kg)', 'Gi√°/kg (VND)', 'T·ªïng chi ph√≠ (VND)'])

        for component, data in sorted(mix_analysis['components_summary'].items(),
                                    key=lambda x: x[1]['total_consumption'], reverse=True):
            overview_sheet.append([
                component,
                round(data['total_consumption'], 2),
                round(data['average_daily'], 2),
                round(data['max_daily'], 2),
                round(data['min_daily'], 2) if data['min_daily'] != float('inf') else 0,
                data['component_info'].get('price_per_kg', 0),
                round(data.get('total_cost', 0), 0)
            ])

        overview_sheet.append([])

        # Th√¥ng tin ch·ª©c nƒÉng
        overview_sheet.append(['TH√îNG TIN CH·ª®C NƒÇNG TH√ÄNH PH·∫¶N'])
        overview_sheet.append(['Th√†nh ph·∫ßn', 'Ch·ª©c nƒÉng', 'Li·ªÅu l∆∞·ª£ng khuy·∫øn ngh·ªã', 'Nh√† cung c·∫•p'])

        for component, data in mix_analysis['components_summary'].items():
            component_info = data['component_info']
            overview_sheet.append([
                component,
                component_info.get('function', 'N/A'),
                component_info.get('dosage_range', 'N/A'),
                component_info.get('supplier', 'N/A')
            ])

        # Apply formatting
        self._apply_advanced_formatting(overview_sheet, f"A1:G{overview_sheet.max_row}", "mix_overview")

    def _create_mix_consumption_trend_sheet(self, workbook: Workbook, mix_analysis: Dict):
        """T·∫°o sheet xu h∆∞·ªõng ti√™u th·ª• mix"""
        trend_sheet = workbook.create_sheet("Xu H∆∞·ªõng Ti√™u Th·ª• Mix")

        # Header
        trend_sheet.append(['XU H∆Ø·ªöNG TI√äU TH·ª§ TH√ÄNH PH·∫¶N MIX THEO NG√ÄY'])
        trend_sheet.append([])

        # T·∫°o b·∫£ng d·ªØ li·ªáu theo ng√†y
        components = list(mix_analysis['components_summary'].keys())
        header = ['Ng√†y'] + components
        trend_sheet.append(header)

        for date_str in sorted(mix_analysis['daily_consumption'].keys()):
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            row = [date_obj.strftime("%d/%m/%Y")]

            daily_data = mix_analysis['daily_consumption'][date_str]
            for component in components:
                amount = daily_data.get(component, 0)
                row.append(round(amount, 2))

            trend_sheet.append(row)

        # Apply formatting
        self._apply_advanced_formatting(trend_sheet, f"A1:F{trend_sheet.max_row}", "mix_trend")

    def _create_mix_efficiency_analysis_sheet(self, workbook: Workbook, mix_analysis: Dict):
        """T·∫°o sheet ph√¢n t√≠ch hi·ªáu qu·∫£ mix"""
        efficiency_sheet = workbook.create_sheet("Ph√¢n T√≠ch Hi·ªáu Qu·∫£ Mix")

        # Header
        efficiency_sheet.append(['PH√ÇN T√çCH HI·ªÜU QU·∫¢ S·ª¨ D·ª§NG TH√ÄNH PH·∫¶N MIX'])
        efficiency_sheet.append([])

        # B·∫£ng ph√¢n t√≠ch hi·ªáu qu·∫£
        efficiency_sheet.append(['Th√†nh ph·∫ßn', 'T·ª∑ l·ªá s·ª≠ d·ª•ng (%)', 'Kho·∫£ng khuy·∫øn ngh·ªã',
                                'Tr·∫°ng th√°i', 'Ghi ch√∫'])

        for component, data in mix_analysis['components_summary'].items():
            usage_rate = data.get('usage_rate', 'N/A')
            dosage_range = data['component_info'].get('dosage_range', 'N/A')
            efficiency_status = data.get('efficiency_status', 'Kh√¥ng x√°c ƒë·ªãnh')

            # Ghi ch√∫ d·ª±a tr√™n tr·∫°ng th√°i
            if efficiency_status == 'D∆∞·ªõi m·ª©c khuy·∫øn ngh·ªã':
                note = 'C·∫ßn tƒÉng li·ªÅu l∆∞·ª£ng'
            elif efficiency_status == 'V∆∞·ª£t m·ª©c khuy·∫øn ngh·ªã':
                note = 'C·∫ßn gi·∫£m li·ªÅu l∆∞·ª£ng'
            elif efficiency_status == 'Trong kho·∫£ng khuy·∫øn ngh·ªã':
                note = 'Li·ªÅu l∆∞·ª£ng ph√π h·ª£p'
            else:
                note = 'C·∫ßn ki·ªÉm tra th√™m'

            efficiency_sheet.append([
                component,
                usage_rate if usage_rate != 'N/A' else 'N/A',
                dosage_range,
                efficiency_status,
                note
            ])

        efficiency_sheet.append([])

        # Khuy·∫øn ngh·ªã t·ªëi ∆∞u h√≥a
        efficiency_sheet.append(['KHUY·∫æN NGH·ªä T·ªêI ∆ØU H√ìA'])
        efficiency_sheet.append(['Th√†nh ph·∫ßn', 'Khuy·∫øn ngh·ªã', 'L√Ω do'])

        for component, data in mix_analysis['components_summary'].items():
            efficiency_status = data.get('efficiency_status', 'Kh√¥ng x√°c ƒë·ªãnh')

            if efficiency_status == 'D∆∞·ªõi m·ª©c khuy·∫øn ngh·ªã':
                recommendation = 'TƒÉng li·ªÅu l∆∞·ª£ng'
                reason = 'Hi·ªáu qu·∫£ ch∆∞a t·ªëi ∆∞u'
            elif efficiency_status == 'V∆∞·ª£t m·ª©c khuy·∫øn ngh·ªã':
                recommendation = 'Gi·∫£m li·ªÅu l∆∞·ª£ng'
                reason = 'Tr√°nh l√£ng ph√≠ v√† t√°c d·ª•ng ph·ª•'
            elif efficiency_status == 'Trong kho·∫£ng khuy·∫øn ngh·ªã':
                recommendation = 'Duy tr√¨ hi·ªán t·∫°i'
                reason = 'Li·ªÅu l∆∞·ª£ng ph√π h·ª£p'
            else:
                recommendation = 'C·∫ßn ƒë√°nh gi√° th√™m'
                reason = 'Thi·∫øu th√¥ng tin tham chi·∫øu'

            efficiency_sheet.append([component, recommendation, reason])

        # Apply formatting
        self._apply_advanced_formatting(efficiency_sheet, f"A1:E{efficiency_sheet.max_row}", "mix_efficiency")


